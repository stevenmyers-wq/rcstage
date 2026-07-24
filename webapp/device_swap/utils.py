import time
import io
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from webapp.rc_api import rc

DEVICE_TYPES = {"SPA-1001": "1", "SPA-3000": "2", "PAP2-NA": "3", "SPA-921": "4", "Other phone": "0", "Softphone": "-1", "PAP2T": "5", "SPIP300": "20", "SPIP320": "22", "SPIP330": "24", "SPIP331": "25", "SPIP430": "26", "SPIP500": "28", "SPIP501": "29", "SPIP550": "30", "SPIP560": "31", "SPIP600": "32", "SPIP601": "33", "SSIP4000": "36", "SSIP6000": "37", "SSIP7000": "38", "VVX1500": "39", "Symbian Phone": "-6", "IPhone": "-5", "BlackBerry": "-7", "Window mobile": "-8", "Android": "-9", "SPA-2102": "7", "Call controller": "-3", "Pagoo": "-2", "SPA-3102": "6", "SPA501G": "13", "SPA502G": "14", "SPA504G": "15", "SPA508G": "16", "SPA-2000": "8", "SPA-942": "11", "SSIP5000": "41", "SPA525G2": "19", "SPA303": "51", "SPA122": "52", "VVX500": "53", "VVX150": "76", "J139": "85", "J179": "86", "B199": "87", "CP-6821-3PCC": "84", "W52P": "57", "CP-8861-3PCC": "74", "CP-7841-3PCC": "73", "Zoom Softphone": "-13", "Zoom Mobile": "-14", "T48S": "80", "W60P": "81", "OBI302": "82", "J169": "88", "BLA main device": "-11", "PagingHangover": "-10", "WebPhone": "-12", "VVX501": "64", "VVX411": "65", "VVX311": "66", "VVX601": "67", "VVX301": "68", "VVX310": "54", "VVX410": "55", "SPA514G": "56", "T21P": "60", "TRIO8500": "75", "J159": "89", "J189": "90", "VVX400": "58", "VVX300": "59", "VVX600": "61", "VVX101": "62", "VVX201": "63", "W56P": "70", "TRIO8800": "69", "T42S": "71", "T46S": "72", "T43U": "135", "T46U": "136", "T48U": "137", "VVXD230": "141", "T31P": "142", "ALEH3G": "138", "ALEH3P": "139", "ALEH6": "140", "VVX250": "77", "SPIP450": "27", "SPIP650": "34", "VVX350": "78", "6920": "155", "6930": "156", "6940": "157", "CP205": "91", "CP400": "92", "CP600": "93", "CP200": "94", "T57W": "95", "CP100": "96", "CP930WB": "102", "SPIP321": "23", "SPIP670": "35", "SPA-941": "10", "CP-6861-3PCC": "158", "ATA191-MPP": "159", "ATA192-MPP": "160", "IP480": "187", "IP480G": "188", "CP110": "184", "CP210": "185", "CP-8841-3PCC": "186", "EDGEE100": "189", "EDGEE300": "190", "CP935WB": "193", "EDGEE500": "192", "J189A": "195", "CP710": "191", "CP410": "194", "ROVE20": "196", "J129": "197", "T34W": "199", "DP-9841": "201", "DP-9851": "202", "DP-9861": "203", "DP-9871": "204", "SPIP301": "21", "AX83H": "205", "AX86R": "206", "RCAIGW": "-17", "T85W": "207", "T73W": "208", "T74W": "209", "CCX505": "200", "T73U": "210", "T74U": "211", "ALE8008": "42", "ALE8008G": "43", "ALE8018": "44", "ALE8058S": "45", "ALE8068S": "46", "ALE8028S": "47", "T40P": "97", "CP700": "98", "SPIP335": "40", "VVX401": "99", "CP920": "103", "T49G": "104", "T19P_E2": "105", "T23G": "106", "T23P": "107", "T27G": "108", "T27P": "109", "T29G": "110", "T40G": "111", "T41P": "112", "T41S": "113", "T42G": "114", "T46G": "115", "T48G": "116", "T52S": "117", "T53": "118", "T53W": "119", "T54S": "120", "T54W": "121", "T56A": "122", "T58": "123", "CP960": "124", "W69P": "125", "CCX400": "126", "CCX500": "127", "CCX600": "128", "CCX700": "129", "TRIOC60": "134", "ALEM3": "130", "ALEM5": "131", "ALEM7": "132", "ALE8078S": "133", "RCV Room": "-15", "6905": "161", "6910": "162", "6970": "163", "6863I": "164", "6865I": "165", "6867I": "166", "6869I": "167", "SPA509G": "17", "SPA-922": "9", "SPA-962": "12", "SPA301": "50", "CP-8851-3PCC": "83", "CP-7821-3PCC": "49", "CP-8811-3PCC": "48", "TRIO8300": "100", "T33G": "101", "RCApp": "-16", "ALE2": "151", "CP700X": "145", "T58WPRO": "146", "W76P": "149", "W79P": "150", "CP925": "147", "CP965": "148", "VP59": "152", "ROVE30": "143", "ROVE40": "144", "POLYEDGEB10": "153", "POLYEDGEB30": "154", "6873I": "168", "SNOM715": "169", "SNOMD717": "170", "SNOM725": "171", "SNOMD735": "172", "SNOMD765": "173", "SNOMD785": "174", "EDGEE220": "178", "EDGEE320": "179", "EDGEE350": "180", "EDGEE450": "181", "EDGEE550": "182", "6920W": "175", "6930W": "176", "6940W": "177", "IP485G": "183", "SPA525G": "18", "VVX450": "79", "T77LTE": "215", "T77U": "212", "T87W": "213", "DT-200A": "214"}
def generate_device_swap_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Devices"
    
    headers = ["Extension", "Device Type", "MAC Address", "Device Name"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
        
    ws_devices = wb.create_sheet(title="DeviceList")
    device_names = sorted(list(DEVICE_TYPES.keys()))
    
    for i, name in enumerate(device_names, start=1):
        ws_devices.cell(row=i, column=1, value=name)
    
    ws_devices.sheet_state = 'hidden'
    
    # openpyxl writes formula1 verbatim into <formula1>, so it must NOT begin
    # with '='. A leading '=' produces invalid XML that makes Excel flag the
    # file as corrupt and drops the dropdown entirely.
    formula = f"DeviceList!$A$1:$A${len(device_names)}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = 'Please select a valid Device Type from the dropdown list.'
    dv.errorTitle = 'Invalid Device'
    
    ws.add_data_validation(dv)
    dv.add('B2:B1000')
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def process_bulk_device_update(records):
    results = []
    devices_to_update = []
    ext_map = {}

    # 1. Map extensionNumber -> extensionId (enabled users only)
    endpoint = '/restapi/v1.0/account/~/extension'
    params = {'perPage': 1000}
    while True:
        resp = rc.get(endpoint, params=params)
        data = resp.json() if hasattr(resp, 'json') else {}
        for ext in data.get('records', []):
            if ext.get('status') == 'Enabled' and ext.get('type') == 'User':
                ext_num = str(ext.get('extensionNumber', ''))
                ext_map[ext_num] = str(ext.get('id'))

        nav = data.get('navigation', {})
        if nav.get('nextPage'):
            params['page'] = params.get('page', 1) + 1
        else:
            break

    # 2. Build one bulk-update record per valid row
    for row in records:
        ext_num = str(row.get('Extension', '')).split('.')[0].strip()
        raw_mac = str(row.get('MAC Address', '')).strip()
        # Strip every non-hex character (colons, dashes, dots, spaces), then lowercase
        target_mac = ''.join(c for c in raw_mac if c in '0123456789abcdefABCDEF').lower()
        device_type_name = str(row.get('Device Type', '')).strip()
        device_name = str(row.get('Device Name', '')).strip()

        if not ext_num or ext_num == 'nan' or not raw_mac or raw_mac == 'nan':
            continue

        result_entry = {'extension': ext_num, 'mac': raw_mac, 'status': 'Failed', 'reason': ''}

        if ext_num not in ext_map:
            result_entry['reason'] = f"Extension {ext_num} not found or not an enabled user."
            results.append(result_entry)
            continue

        model_id = DEVICE_TYPES.get(device_type_name)
        if not model_id:
            result_entry['reason'] = f"Invalid Device Type '{device_type_name}'."
            results.append(result_entry)
            continue

        try:
            ext_id = ext_map[ext_num]
            dev_resp = rc.get(f'/restapi/v1.0/account/~/extension/{ext_id}/device')
            dev_data = dev_resp.json() if hasattr(dev_resp, 'json') else {}
            dev_records = dev_data.get('records', [])

            # Prefer a physical phone: HardPhone -> OtherPhone -> SoftPhone.
            # (Grabbing whatever came first could target a RingCentral App softphone
            #  and fail; this ordering matches the known-good tool.)
            target_device = (
                next((d for d in dev_records if d.get('type') == 'HardPhone'), None)
                or next((d for d in dev_records if d.get('type') == 'OtherPhone'), None)
                or next((d for d in dev_records if d.get('type') == 'SoftPhone'), None)
            )

            if not target_device:
                result_entry['reason'] = "No eligible device (HardPhone/OtherPhone/SoftPhone) on this extension."
                results.append(result_entry)
                continue

            device_id = str(target_device['id'])

            # RingCentral's bulk-update identifies each device by "deviceId".
            update_obj = {
                "deviceId": device_id,
                "serial": target_mac,
                "model": {"id": str(model_id)}
            }
            if device_name and device_name.lower() != 'nan':
                update_obj['name'] = device_name

            devices_to_update.append(update_obj)
            result_entry['status'] = 'Pending'
            result_entry['device_id'] = device_id
            results.append(result_entry)

        except Exception as e:
            result_entry['reason'] = f"Failed to look up device: {str(e)}"
            results.append(result_entry)

    # 3. Fire a single bulk-update request
    #    POST /restapi/v1.0/account/~/device/bulk-update  ->  {"records": [...]}
    if not devices_to_update:
        return results

    try:
        bulk_resp = rc.post('/restapi/v1.0/account/~/device/bulk-update',
                            json={"records": devices_to_update})
        status_code = getattr(bulk_resp, 'status_code', 500)

        # Whole request rejected (auth, malformed, etc.) -> fail all pending rows
        if status_code >= 400:
            error_msg = f"HTTP {status_code}"
            try:
                err_json = bulk_resp.json()
                error_msg = err_json.get('message') or err_json.get('description') or error_msg
            except Exception:
                pass
            for r in results:
                if r.get('status') == 'Pending':
                    r['status'] = 'Failed'
                    r['reason'] = f"API Error: {error_msg}"
            return _clean_results(results)

        bulk_data = bulk_resp.json() if hasattr(bulk_resp, 'json') else {}
        returned = bulk_data.get('records', [])
        pending = [r for r in results if r.get('status') == 'Pending']

        # Map each returned record back to its row by deviceId, falling back to order
        for idx, api_rec in enumerate(returned):
            dev_id = str(api_rec.get('deviceId') or api_rec.get('id') or '')
            match = next((r for r in pending if r.get('device_id') == dev_id), None)
            if not match and idx < len(pending):
                match = pending[idx]
            if not match:
                continue

            if api_rec.get('successful'):
                match['status'] = 'Success'
                match['reason'] = 'Device updated successfully.'
            else:
                err = api_rec.get('error') or {}
                msg = err.get('message') or err.get('description') or 'Update rejected by RingCentral.'
                code = err.get('errorCode', 'N/A')
                match['status'] = 'Failed'
                match['reason'] = f"API Error: {msg} (Code: {code})"

        # Any pending row with no corresponding response record
        for r in results:
            if r.get('status') == 'Pending':
                r['status'] = 'Failed'
                r['reason'] = 'No response received from RingCentral for this device.'

    except Exception as e:
        for r in results:
            if r.get('status') == 'Pending':
                r['status'] = 'Failed'
                r['reason'] = f"Execution error: {str(e)}"

    return _clean_results(results)


def _clean_results(results):
    """Drop internal bookkeeping keys before returning to the frontend."""
    for r in results:
        r.pop('device_id', None)
    return results


# ---------------------------------------------------------------------------
# Bulk Device Swap  (RingCentral v2 "replace" — swap one existing device for
# another existing device: from inventory or from another extension).
# Different operation from the Device Update above, which re-provisions the
# model/serial of the device already on an extension.
# ---------------------------------------------------------------------------

# RingCentral only swaps these device types via the /replace endpoint.
SWAPPABLE_TYPES = ('HardPhone', 'OtherPhone')


def _norm_mac(value):
    """Normalise a MAC/serial to lowercase hex only (strip : - . spaces)."""
    return ''.join(c for c in str(value or '') if c in '0123456789abcdefABCDEF').lower()


def generate_swap_template():
    """Blank template for the Bulk Swap tab (no device-type dropdown needed)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Swaps"

    headers = ["Extension", "Replacement Device", "Current Device"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col, width in {'A': 16, 'B': 34, 'C': 34}.items():
        ws.column_dimensions[col].width = width

    info = wb.create_sheet(title="Instructions")
    for line in [
        ["Bulk Device Swap — how to fill this in"],
        [""],
        ["Each row swaps the device currently on an Extension for the Replacement Device."],
        ["The phone number, digital line and emergency address stay on the line and move to the replacement."],
        ["The device that was on the extension is returned to inventory (unassigned)."],
        ["RingCentral can only swap HardPhone and 'Existing phone' (OtherPhone) devices."],
        [""],
        ["Extension           (required):  the extension that keeps its number and receives the replacement device."],
        ["Replacement Device  (required):  MAC or device Name of the device to install."],
        ["                                 May be an unassigned/inventory phone, or a device on another extension."],
        ["Current Device      (optional):  MAC or Name of the device currently on the Extension to swap out."],
        ["                                 Only needed if the extension has more than one device."],
    ]:
        info.append(line)
    info.column_dimensions['A'].width = 115

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _fetch_all_records(path_base):
    """Paginate a v1.0 list endpoint and return all records."""
    out = []
    page = 1
    while True:
        sep = '&' if '?' in path_base else '?'
        resp = rc.get(f"{path_base}{sep}perPage=1000&page={page}")
        data = resp.json() if hasattr(resp, 'json') else {}
        recs = data.get('records', [])
        out.extend(recs)
        if not data.get('navigation', {}).get('nextPage'):
            break
        page += 1
        time.sleep(0.05)
    return out


def process_device_swaps(records):
    results = []

    # 1. Extension number <-> id maps
    ext_num_to_id = {}
    for e in _fetch_all_records('/restapi/v1.0/account/~/extension'):
        num = str(e.get('extensionNumber', '')).strip()
        eid = str(e.get('id', '')).strip()
        if num and eid:
            ext_num_to_id[num] = eid
    ext_id_to_num = {v: k for k, v in ext_num_to_id.items()}

    # 2. All account devices (assigned + inventory)
    devices = _fetch_all_records('/restapi/v1.0/account/~/device')

    def dev_ext_id(d):
        ext = d.get('extension') or {}
        return str(ext.get('id')) if ext.get('id') else None

    def matches(d, identifier):
        ident = str(identifier or '').strip()
        if not ident:
            return False
        mac = _norm_mac(ident)
        serial = _norm_mac(d.get('serial', ''))
        name = str(d.get('name', '')).strip().lower()
        return (mac and serial and mac == serial) or (name == ident.lower())

    def describe(d):
        model = (d.get('model') or {}).get('name') or d.get('type', 'Device')
        ident = d.get('serial') or d.get('name') or str(d.get('id'))
        return f"{model} · {ident}"

    def where(d):
        eid = dev_ext_id(d)
        return f"Ext {ext_id_to_num.get(eid, eid)}" if eid else "inventory"

    for row in records:
        ext_num = str(row.get('Extension', '')).split('.')[0].strip()
        replacement = str(row.get('Replacement Device', '')).strip()
        current = str(row.get('Current Device', '')).strip()

        ext_num = '' if ext_num.lower() == 'nan' else ext_num
        replacement = '' if replacement.lower() == 'nan' else replacement
        current = '' if current.lower() == 'nan' else current

        # Skip blank rows entirely
        if not ext_num and not replacement and not current:
            continue

        entry = {
            'extension': f"Ext {ext_num}" if ext_num else '',
            'replacement': replacement,
            'status': 'Failed',
            'reason': ''
        }

        # --- Extension (required): the line that keeps its number ---
        if not ext_num:
            entry['reason'] = "Extension is required."
            results.append(entry)
            continue
        if ext_num not in ext_num_to_id:
            entry['reason'] = f"Extension {ext_num} not found."
            results.append(entry)
            continue
        ext_id = ext_num_to_id[ext_num]

        # --- Device currently on that extension (the one being swapped out) ---
        on_ext = [d for d in devices if dev_ext_id(d) == ext_id]
        if current:
            outgoing_matches = [d for d in on_ext if matches(d, current)]
        else:
            outgoing_matches = on_ext

        if not outgoing_matches:
            if current:
                entry['reason'] = f"No device on extension {ext_num} matching '{current}'."
            else:
                entry['reason'] = f"Extension {ext_num} has no device to swap out."
            results.append(entry)
            continue
        if len(outgoing_matches) > 1:
            entry['reason'] = (f"Extension {ext_num} has {len(outgoing_matches)} devices — "
                               f"put a MAC or name in 'Current Device' to pick which to replace.")
            results.append(entry)
            continue
        outgoing = outgoing_matches[0]

        # --- Replacement device (required): found anywhere on the account ---
        if not replacement:
            entry['reason'] = "Replacement Device is required."
            results.append(entry)
            continue
        incoming_matches = [d for d in devices if matches(d, replacement)]
        if not incoming_matches:
            entry['reason'] = f"Replacement device '{replacement}' not found on this account."
            results.append(entry)
            continue
        if len(incoming_matches) > 1:
            entry['reason'] = (f"Replacement '{replacement}' matches {len(incoming_matches)} devices — "
                               f"use the MAC to identify it uniquely.")
            results.append(entry)
            continue
        incoming = incoming_matches[0]

        outgoing_id = str(outgoing.get('id'))
        incoming_id = str(incoming.get('id'))
        entry['extension'] = f"Ext {ext_num} · out: {describe(outgoing)}"
        entry['replacement'] = f"{describe(incoming)} (from {where(incoming)})"

        if outgoing_id == incoming_id:
            entry['reason'] = "The replacement device is already on this extension."
            results.append(entry)
            continue

        # --- Execute: replace the outgoing device (on the extension) with the incoming one ---
        #     POST /restapi/v2/accounts/~/extensions/{extId}/devices/{outgoingId}/replace
        try:
            endpoint = f"/restapi/v2/accounts/~/extensions/{ext_id}/devices/{outgoing_id}/replace"
            resp = rc.post(endpoint, json={"targetDeviceId": incoming_id})
            code = getattr(resp, 'status_code', 500)

            if code in (200, 201, 202, 204):
                entry['status'] = 'Success'
                entry['reason'] = 'Device swapped successfully.'
            else:
                msg = f"HTTP {code}"
                try:
                    j = resp.json()
                    if isinstance(j, dict):
                        errs = j.get('errors') or []
                        if errs and isinstance(errs[0], dict):
                            msg = errs[0].get('message') or msg
                            if errs[0].get('errorCode'):
                                msg = f"{msg} (Code: {errs[0]['errorCode']})"
                        else:
                            msg = j.get('message') or j.get('description') or msg
                except Exception:
                    pass
                entry['status'] = 'Failed'
                entry['reason'] = f"API Error: {msg}"
        except Exception as e:
            entry['status'] = 'Failed'
            entry['reason'] = f"Execution error: {str(e)}"

        results.append(entry)

    return results
