import os
import re
import copy
import time
import io
import json
import pandas as pd
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from webapp.rc_api import rc_api_call
from webapp import task_control

audit_progress_store = {}

# Set the CQ_DEBUG env var (1/true/yes/on) to have every failed queue/answering-rule
# write dump the exact JSON payload we sent plus RingCentral's raw error into the
# System Log. This is how the less-documented bits (action enums, fixedOrderAgents
# shape) can be dialed in against a live account without guessing.
CQ_DEBUG = str(os.environ.get('CQ_DEBUG', '')).lower() in ('1', 'true', 'yes', 'on')

# Service-web host used to build browser-downloadable greeting links. The API's platform
# host (platform.ringcentral.com) is global even for AU accounts, so the console region
# can't be derived from it — configure via env var; default matches this deployment (AU).
RC_SERVICE_WEB_HOST = os.environ.get('RC_SERVICE_WEB_HOST', 'service.ringcentral.com.au').strip()


def _greeting_download_url(custom_obj):
    """Build a browser-downloadable greeting URL (works with a logged-in service-web
    session) from a custom greeting's {id, uri}. Returns '' if it can't be built."""
    if not isinstance(custom_obj, dict):
        return ''
    g_id = str(custom_obj.get('id', '')).strip()
    m = re.search(r'/extension/(\d+)', str(custom_obj.get('uri', '')))
    mailbox_id = m.group(1) if m else ''
    if g_id and mailbox_id:
        return f"https://{RC_SERVICE_WEB_HOST}/mobile/media?cmd=downloadGreeting&gId={g_id}&mailboxId={mailbox_id}"
    return ''


def _debug_dump(logs, label, payload=None, err=None):
    """Append a payload/error snapshot to a row's log list when CQ_DEBUG is on, and
    mirror it to stdout so it also lands in the Cloud Run logs."""
    if not CQ_DEBUG:
        return
    parts = [f"[DEBUG] {label}"]
    if payload is not None:
        try:
            parts.append(f"payload={json.dumps(payload, default=str)[:1800]}")
        except Exception:
            parts.append("payload=<unserializable>")
    if err is not None:
        parts.append(f"raw_error={str(err)[:1200]}")
    line = " | ".join(parts)
    logs.append(line)
    print(f"[CQ_DEBUG] {line}", flush=True)

DAY_ABBR = {
    "mon": "monday", "tue": "tuesday", "wed": "wednesday",
    "thu": "thursday", "fri": "friday", "sat": "saturday", "sun": "sunday"
}

_READ_ONLY = ('uri', 'id', 'type', 'name', 'creationTime', 'lastModifiedTime')

# Genre names valid for the music slots (Audio While Connecting / Hold Music).
AUDIO_GENRES = frozenset({
    'ring tones', 'acoustic', 'beautiful', 'classical', 'corporate', 'country',
    'electronic', 'modern jazz', 'nature', 'pop', 'r&b', 'rock', 'upbeat'
})

# --- Shared workbook schema (used by both the download template and the audit export) ---
# Kept in one place so the two exports can never drift apart.
TEMPLATE_COLUMNS = [
    "Queue Name", "Record Group Name", "Extension", "Site", "Status", "Phone Number",
    "Queue Manager", "Queue Email", "Queue PIN", "Members (Ext)", "Timezone", "Hours",
    "Greeting", "Audio While Connecting", "Hold Music", "Interrupt Audio", "Interrupt Prompt",
    "Ring Type", "User Ring Time", "Total Ring Time", "Wrap Up Time", "Member Queue Status",
    "Callers In Queue", "When Queue is Full", "Queue Full Destination", "When Max Time is Reached",
    "Time Reached Destination", "Voicemail Greeting", "Voicemail Recipients",
    "Voicemail Notifications", "Voicemail Notifications Email", "After Hours Behavior",
    "After Hours Destination", "Voicemail to Text", "Missed Call Notifications",
    "Inbound Fax Notifications", "Outbound Fax Notifications", "Text Notifications",
    "Missed Call Notifications Email", "Inbound Fax Notifications Email",
    "Outbound Fax Notifications Email", "Text Notifications Email"
]

# Per-type notification columns: (sheet On/Off toggle column, sheet per-type email column,
# RingCentral notification-settings key). Voicemail is handled separately (it has its own
# style column). Per-type email addresses are an advanced-mode feature.
NOTIF_TYPE_COLS = [
    ('Missed Call Notifications', 'Missed Call Notifications Email', 'missedCalls'),
    ('Inbound Fax Notifications', 'Inbound Fax Notifications Email', 'inboundFaxes'),
    ('Outbound Fax Notifications', 'Outbound Fax Notifications Email', 'outboundFaxes'),
    ('Text Notifications', 'Text Notifications Email', 'inboundTexts'),
]

GLOBAL_TIMEZONES = [
    "US/Eastern", "US/Central", "US/Mountain", "US/Pacific", "US/Alaska", "US/Hawaii",
    "Canada/Eastern", "Canada/Central", "Canada/Mountain", "Canada/Pacific", "Canada/Atlantic",
    "Europe/London", "Europe/Paris", "Europe/Berlin", "Europe/Athens", "Europe/Moscow",
    "GMT", "UTC", "Asia/Dubai", "Asia/Kolkata", "Asia/Singapore", "Asia/Tokyo", "Asia/Hong_Kong",
    "Australia/Sydney", "Australia/Melbourne", "Australia/Brisbane", "Australia/Adelaide", "Australia/Perth",
    "Pacific/Auckland", "America/Sao_Paulo", "America/Buenos_Aires", "America/Mexico_City"
]

# Column letter -> Excel list-validation formula. Letters map to TEMPLATE_COLUMNS positions.
SCHEMA_VALIDATIONS = {
    "E": '"Enabled,Disabled"',
    "M": '"Default,Custom,Off"',
    "N": '"Default,Ring Tones,Acoustic,Beautiful,Classical,Corporate,Country,Electronic,Modern Jazz,Nature,Pop,R&B,Rock,Upbeat,Custom,Off"',
    "O": '"Default,Ring Tones,Acoustic,Beautiful,Classical,Corporate,Country,Electronic,Modern Jazz,Nature,Pop,R&B,Rock,Upbeat,Custom,Off"',
    "P": '"Never,10 Seconds,15 Seconds,20 Seconds,25 Seconds,30 Seconds,40 Seconds,50 Seconds,1 Minute"',
    "Q": '"Thank you for your patience,Higher than normal volume,Agents are currently busy,Call is very important to us,Custom,Default,Off"',
    "R": '"Simultaneous,Sequential,Rotating"',
    "S": '"10 Seconds,15 Seconds,20 Seconds,25 Seconds,30 Seconds,40 Seconds,50 Seconds,1 Minute,2 Minutes"',
    "T": '"15 Seconds,30 Seconds,45 Seconds,1 Minute,2 Minutes,3 Minutes,4 Minutes,5 Minutes,10 Minutes,15 Minutes"',
    "U": '"0 Seconds,5 Seconds,10 Seconds,15 Seconds,20 Seconds,30 Seconds,1 Minute"',
    "V": '"Allowed,Not Allowed"',
    "W": '"5,10,15,20,25"',
    "X": '"Voicemail,TransferToExtension,Disconnect,Announcement"',
    "Z": '"Voicemail,TransferToExtension,Disconnect,Announcement"',
    "AB": '"Default,Custom,Off"',
    "AD": '"Off,Notify by Email,Notify & Attach,Notify Attach & Read"',
    "AF": '"TakeMessagesOnly,TransferToExtension,UnconditionalForwarding,PlayAnnouncementOnly,Disconnect"',
    "AH": '"On,Off"',
    "AI": '"On,Off"',
    "AJ": '"On,Off"',
    "AK": '"On,Off"',
    "AL": '"On,Off"'
}


def build_config_workbook(df):
    """Render a Queue Config DataFrame to an .xlsx BytesIO with the timezone
    reference sheet, dropdown validations and text number-formatting applied.
    Shared by the download-template and audit-export paths."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Queue Config')

        tz_df = pd.DataFrame({"ValidTimezones": GLOBAL_TIMEZONES})
        tz_df.to_excel(writer, index=False, sheet_name='Timezone_Ref')

        workbook = writer.book
        config_ws = workbook['Queue Config']

        dv_tz = DataValidation(
            type="list",
            formula1="=Timezone_Ref!$A$2:$A$" + str(len(GLOBAL_TIMEZONES) + 1),
            allow_blank=True
        )
        config_ws.add_data_validation(dv_tz)
        dv_tz.add("K2:K1000")

        for col_letter, formula_string in SCHEMA_VALIDATIONS.items():
            dv = DataValidation(type="list", formula1=formula_string, allow_blank=True)
            config_ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

        # Force the whole config grid to Text so Excel doesn't auto-convert entries
        # (24/7 -> a date, extensions losing leading zeros, long numbers -> scientific).
        # openpyxl only formats cells that exist, so set the column default AND stamp every
        # row the user might type into (not just the written header/data rows).
        n_cols = df.shape[1]
        for c in range(1, n_cols + 1):
            col_letter = get_column_letter(c)
            config_ws.column_dimensions[col_letter].number_format = '@'
            for r in range(1, 1001):
                config_ws.cell(row=r, column=c).number_format = '@'

    output.seek(0)
    return output

def to_int(val):
    try: return int(float(val))
    except (TypeError, ValueError): return None

def _parse_toggle(val):
    """Interpret an on/off notification cell. Returns True or False for a recognized
    value, or None when the cell is blank/unrecognized so the caller leaves the
    corresponding setting untouched."""
    if val is None: return None
    v = str(val).strip().lower()
    if v in ('on', 'enabled', 'enable', 'true', 'yes', '1', 'y'): return True
    if v in ('off', 'disabled', 'disable', 'false', 'no', '0', 'n'): return False
    return None

def parse_time_to_seconds(val):
    if pd.isna(val) or val == '': return None
    val_str = str(val).lower().strip()
    match = re.search(r'([\d.]+)', val_str)
    if not match: return None
    num = float(match.group(1))
    return int(num * 60) if 'min' in val_str else int(num)

def format_sec(val):
    if val is None or str(val).strip() == "" or str(val).strip() == "None": return "None"
    try:
        v = int(float(val))
        if v == 0: return "0 Seconds"
        if v >= 60 and v % 60 == 0:
            mins = v // 60
            return f"{mins} Minute{'s' if mins > 1 else ''}"
        return f"{v} Seconds"
    except: return str(val)

def format_schedule(schedule_dict):
    if not schedule_dict: return "24/7"
    day_order = {"monday": 1, "tuesday": 2, "wednesday": 3, "thursday": 4, "friday": 5, "saturday": 6, "sunday": 7}
    sorted_days = sorted(schedule_dict.items(), key=lambda x: day_order.get(x[0].lower(), 99))
    
    days = []
    for day, times in sorted_days:
        time_strs = []
        for t in times:
            if t['from'] == "00:00" and t['to'] == "23:59": time_strs.append("24/7")
            else: time_strs.append(f"{t['from']}-{t['to']}")
        days.append(f"{day[:3].capitalize()}: {', '.join(time_strs)}")
    return " | ".join(days)

def to_24h(time_str):
    time_str = time_str.replace(" ", "").lower()
    if "am" in time_str or "pm" in time_str:
        fmt = "%I:%M%p" if ":" in time_str else "%I%p"
        return datetime.strptime(time_str, fmt).strftime("%H:%M")
    return datetime.strptime(time_str, "%H:%M").strftime("%H:%M")

def parse_intuitive_hours(hours_str):
    hours_str = str(hours_str).lower().strip().replace("\n", " ").replace("–", "-").replace("—", "-")
    hours_str = re.sub(r'(?<!\d)(\d{1,2})\.(\d{2})\s*([ap]m)?', r'\1:\2\3', hours_str)
    hours_str = re.sub(r'(?<!\d)(\d{1,2}(?::\d{2})?)\s*(?:-|to|thru|through)\s*(\d{1,2}(?::\d{2})?\s*pm)', r'\1am-\2', hours_str)
    
    if hours_str in ['24/7', '24x7', '24-7', '24 7']: return "24/7"
    if not hours_str or hours_str in ['closed', 'none', 'n/a', 'off']: return {} 
        
    time_pattern = r'(?:\d{1,2}(?::\d{2})?\s*[ap]m|\d{1,2}:\d{2})\s*(?:-|to|thru|through)\s*(?:\d{1,2}(?::\d{2})?\s*[ap]m|\d{1,2}:\d{2})'
    parts = re.split(f'({time_pattern})', hours_str)
    if len(parts) == 1: raise ValueError(f"Could not detect valid time ranges in text: '{hours_str}'.")
        
    times = [parts[i] for i in range(1, len(parts), 2)]
    texts = [parts[i] for i in range(0, len(parts), 2)]
    
    day_map = {
        r'\bmonday\b': 'mon', r'\bm\b': 'mon', r'\btuesday\b': 'tue', r'\btues\b': 'tue', r'\btu\b': 'tue',
        r'\bwednesday\b': 'wed', r'\bw\b': 'wed', r'\bthursday\b': 'thu', r'\bthurs\b': 'thu', r'\bthu\b': 'thu', r'\bth\b': 'thu',
        r'\bfriday\b': 'fri', r'\bf\b': 'fri', r'\bsaturday\b': 'sat', r'\bsa\b': 'sat', r'\bsunday\b': 'sun', r'\bsu\b': 'sun'
    }
    
    days_before = False
    if texts and texts[0].strip():
        days_before = bool(re.search(r'\b(m|tu|tue|tues|w|wed|th|thu|thurs|f|fri|sa|sat|su|sun|mon|monday|tuesday|wednesday|thursday|friday|saturday|sunday)\b', texts[0]))

    weekly_ranges = {}
    keys = list(DAY_ABBR.keys())
    
    for i, t in enumerate(times):
        assoc_text = texts[i] if days_before else texts[i+1]
        for pat, rep in day_map.items(): assoc_text = re.sub(pat, rep, assoc_text)
            
        time_match = re.search(r'(\d{1,2}(?::\d{2})?\s*[ap]m|\d{1,2}:\d{2})\s*(?:-|to|thru|through)\s*(\d{1,2}(?::\d{2})?\s*[ap]m|\d{1,2}:\d{2})', t)
        if not time_match: continue
        start_24 = to_24h(time_match.group(1))
        end_24 = to_24h(time_match.group(2))
        
        day_tokens = re.findall(r'(mon|tue|wed|thu|fri|sat|sun)', assoc_text)
        day_ranges = re.findall(r'(mon|tue|wed|thu|fri|sat|sun)\s*(?:-|to|thru|through)\s*(mon|tue|wed|thu|fri|sat|sun)', assoc_text)
        
        days_to_apply = set()
        if day_ranges:
            for d_start, d_end in day_ranges:
                idx_start, idx_end = keys.index(d_start), keys.index(d_end)
                if idx_start <= idx_end:
                    for j in range(idx_start, idx_end + 1): days_to_apply.add(keys[j])
                else: 
                    for j in range(idx_start, 7): days_to_apply.add(keys[j])
                    for j in range(0, idx_end + 1): days_to_apply.add(keys[j])
        elif day_tokens:
            for dt in day_tokens: days_to_apply.add(dt)
        else:
            if i == 0 and len(times) == 1: days_to_apply = set(keys)
            
        for d in days_to_apply:
            full_day = DAY_ABBR[d]
            if full_day not in weekly_ranges: weekly_ranges[full_day] = []
            weekly_ranges[full_day].append({"from": start_24, "to": end_24})
            
    if not weekly_ranges: raise ValueError("Could not detect any valid days associated with the provided times.")
    return weekly_ranges

def format_api_error(err_str):
    if not err_str or str(err_str).strip() == "": 
        return "Unknown API Error (Empty Response)"
    try:
        err_json = json.loads(err_str)
        if isinstance(err_json, dict):
            if 'errors' in err_json and err_json['errors']:
                msgs = []
                for e in err_json['errors']:
                    code = e.get('errorCode', 'Error')
                    msg = e.get('message', '')
                    param = e.get('parameterName', '')
                    if param: msgs.append(f"{code}: {msg} [{param}]")
                    else: msgs.append(f"{code}: {msg}")
                return " | ".join(msgs)
            msg = err_json.get('message', '')
            if msg: return msg
        return str(err_str)
    except:
        return str(err_str)

def safe_api_call(endpoint, method='GET', json_payload=None, token=None, max_retries=4):
    if 'mock_' in str(endpoint):
        if method == 'GET':
            if 'business-hours' in str(endpoint): return True, {"schedule": {"weeklyRanges": {}}}
            if 'answering-rule' in str(endpoint): return True, {"queue": {}}
            if 'notification-settings' in str(endpoint): return True, {"voicemails": {}}
            if 'members' in str(endpoint): return True, {"records": []}
            if 'managers' in str(endpoint): return True, {"records": []}
            return True, {"name": "New Queue", "status": "NotActivated", "contact": {}}
        return True, {}

    for attempt in range(max_retries):
        try:
            resp = rc_api_call(endpoint, method=method, json=json_payload, token=token, return_response=True)
            status_code = getattr(resp, 'status_code', None)
            if status_code == 429:
                try: retry_after = int(resp.headers.get('Retry-After', 60))
                except: retry_after = 60
                time.sleep(retry_after + 1)
                continue
            if resp and getattr(resp, 'ok', False):
                try: return True, resp.json() if resp.content else {}
                except: return True, {}
            try: 
                err_dict = resp.json()
                err_msg = json.dumps(err_dict)
            except: 
                body_text = getattr(resp, 'text', '')
                err_msg = body_text if body_text else f'HTTP {status_code} Error (empty response body)'
            return False, err_msg
        except Exception as e: 
            time.sleep(2)
    return False, "Max retries exceeded due to rate limiting."

def fetch_directory(endpoint, token):
    records = []
    page = 1
    while True:
        sep = "&" if "?" in endpoint else "?"
        succ, resp = safe_api_call(f'{endpoint}{sep}perPage=1000&page={page}', method='GET', token=token)
        if not succ:
            return False, resp
        if isinstance(resp, dict) and 'records' in resp:
            records.extend(resp['records'])
            if 'navigation' in resp and 'nextPage' in resp.get('navigation', {}): page += 1
            else: break
        else:
            break
    return True, records

def fetch_all_queues(token):
    """Fetches all call queues and formats them for the UI."""
    succ, records = fetch_directory('/restapi/v1.0/account/~/call-queues', token)
    if not succ:
        raise Exception(f"Failed to fetch call queues: {format_api_error(records)}")
        
    queues = []
    for q in records:
        queues.append({
            "id": str(q.get('id', '')),
            "name": q.get('name', 'Unknown'),
            "extensionNumber": str(q.get('extensionNumber', '')),
            "site": q.get('site', {}).get('name', 'Main Site')
        })
    
    # Sort alphabetically by name
    return sorted(queues, key=lambda x: x['name'].lower())

def _safe_get_transfer_id(transfer_data, action_type):
    if not transfer_data: return ''
    if isinstance(transfer_data, list):
        for item in transfer_data:
            if item.get('action') == action_type:
                return str(item.get('extension', {}).get('id', ''))
    return ''

def _set_queue_transfer(q_set, action_type, ext_id):
    if 'transfer' not in q_set or not isinstance(q_set['transfer'], list):
        q_set['transfer'] = []
    
    q_set['transfer'] = [t for t in q_set['transfer'] if t.get('action') != action_type]
    
    if ext_id and ext_id != 'None':
        q_set['transfer'].append({
            "extension": {"id": ext_id},
            "action": action_type
        })
        
    if len(q_set['transfer']) == 0:
        q_set.pop('transfer', None)

def _safe_get_ah_transfer_id(transfer_data):
    if not transfer_data: return ''
    # RingCentral returns the answering-rule transfer as an object
    # ({"extension": {"id": ...}}); tolerate the list form too for safety.
    if isinstance(transfer_data, list) and len(transfer_data) > 0:
        return str(transfer_data[0].get('extension', {}).get('id', ''))
    if isinstance(transfer_data, dict):
        return str(transfer_data.get('extension', {}).get('id', ''))
    return ''

def get_old_greeting_name(orig_rule, slot_type):
    for g in orig_rule.get('greetings', []):
        if g.get('type') == slot_type:
            if 'preset' in g:
                name = g['preset'].get('name', 'Default')
                if name.lower() == 'none': return 'Off'
                return name
            elif 'custom' in g:
                return 'Custom'
    return 'Default'

def run_cq_audit(task_id, queue_ids, token):
    audit_progress_store[task_id] = {'current': 0, 'total': len(queue_ids), 'status': 'running', 'file_ready': False}
    try:
        ext_id_to_num = {}
        succ, ext_records = fetch_directory('/restapi/v1.0/account/~/extension', token)
        if succ:
            for e in ext_records:
                ext_id_to_num[str(e['id'])] = str(e.get('extensionNumber', ''))

        succ, sites_resp = safe_api_call('/restapi/v1.0/account/~/sites', token=token)
        site_map = {str(s['id']): s['name'] for s in sites_resp.get('records', [])} if succ else {}

        succ, tz_resp = fetch_directory('/restapi/v1.0/dictionary/timezone', token)
        tz_map = {str(t['id']): t['name'] for t in tz_resp} if succ else {}

        preset_dict = {'Introductory': {}, 'ConnectingAudio': {}, 'HoldMusic': {}, 'InterruptPrompt': {}, 'Voicemail': {}}
        preset_id_to_name = {'Introductory': {}, 'ConnectingAudio': {}, 'HoldMusic': {}, 'InterruptPrompt': {}, 'Voicemail': {}}
        
        # Two-part fetch ensures we get both answering rule defaults AND global presets
        for g_type in preset_dict.keys():
            succ1, dict1 = safe_api_call(f'/restapi/v1.0/dictionary/greeting?greetingType={g_type}&usageType=DepartmentExtensionAnsweringRule&perPage=1000', method='GET', token=token)
            succ2, dict2 = safe_api_call(f'/restapi/v1.0/dictionary/greeting?greetingType={g_type}&perPage=1000', method='GET', token=token)
            
            for dict_resp in [dict1, dict2]:
                if dict_resp and isinstance(dict_resp, dict) and 'records' in dict_resp:
                    for rec in dict_resp['records']:
                        rec_type = rec.get('type')
                        if rec_type and rec_type != g_type:
                            continue 
                        k = rec.get('name', '').lower().strip()
                        v = str(rec.get('id', ''))
                        preset_dict[g_type][k] = v
                        preset_id_to_name[g_type][v] = str(rec.get('name', '')).title()

        rows = []
        for idx, qid in enumerate(queue_ids):
            audit_progress_store[task_id]['current'] = idx + 1
            row = {}
            fixed_order_nums = []

            succ, base = safe_api_call(f'/restapi/v1.0/account/~/extension/{qid}', token=token)
            if not succ: continue
            
            row["Queue Name"] = base.get('name', '')
            row["Extension"] = base.get('extensionNumber', '')
            row["Status"] = base.get('status', '').capitalize()
            row["Queue Email"] = base.get('contact', {}).get('email', '')
            
            # editableMemberStatus lives on the call-queue object, not the extension.
            succ_cq, cq_base = safe_api_call(f'/restapi/v1.0/account/~/call-queues/{qid}', token=token)
            editable = cq_base.get('editableMemberStatus') if succ_cq and isinstance(cq_base, dict) else None
            if editable is True:
                row["Member Queue Status"] = "Allowed"
            elif editable is False:
                row["Member Queue Status"] = "Not Allowed"
            
            site_id = str(base.get('site', {}).get('id', ''))
            row["Site"] = site_map.get(site_id, site_id) if site_id else 'Main Site'
            
            tz_id = str(base.get('regionalSettings', {}).get('timezone', {}).get('id', ''))
            row["Timezone"] = tz_map.get(tz_id, tz_id)

            succ, bh = safe_api_call(f'/restapi/v1.0/account/~/extension/{qid}/business-hours', token=token)
            if succ and bh.get('schedule', {}).get('weeklyRanges'):
                row["Hours"] = format_schedule(bh['schedule']['weeklyRanges'])
            else:
                row["Hours"] = "24/7"

            succ, rule = safe_api_call(f'/restapi/v1.0/account/~/extension/{qid}/answering-rule/business-hours-rule', token=token)
            if succ:
                q_set = rule.get('queue', {})
                
                transfer_mode = q_set.get('transferMode', 'Simultaneous')
                if transfer_mode == 'FixedOrder': row["Ring Type"] = "Sequential"
                elif transfer_mode == 'Rotating': row["Ring Type"] = "Rotating"
                else: row["Ring Type"] = transfer_mode

                # For Sequential queues the agent ring order matters, so remember it and export
                # Members in that exact order (below) — otherwise a plain re-upload would scramble it.
                if transfer_mode == 'FixedOrder':
                    fixed_order_nums = [ext_id_to_num.get(str(a.get('extension', {}).get('id', '')), '')
                                        for a in q_set.get('fixedOrderAgents', [])]
                    fixed_order_nums = [n for n in fixed_order_nums if n]
                
                row["User Ring Time"] = format_sec(q_set.get('agentTimeout'))
                row["Total Ring Time"] = format_sec(q_set.get('holdTime'))
                row["Wrap Up Time"] = format_sec(q_set.get('wrapUpTime'))
                row["Callers In Queue"] = q_set.get('maxCallers')
                
                row["When Queue is Full"] = q_set.get('maxCallersAction')
                f_ext = _safe_get_transfer_id(q_set.get('transfer'), 'MaxCallers')
                if f_ext: row["Queue Full Destination"] = ext_id_to_num.get(f_ext, f_ext)

                row["When Max Time is Reached"] = q_set.get('holdTimeExpirationAction')
                t_ext = _safe_get_transfer_id(q_set.get('transfer'), 'HoldTimeExpiration')
                if t_ext: row["Time Reached Destination"] = ext_id_to_num.get(t_ext, t_ext)

                mode = q_set.get('holdAudioInterruptionMode')
                if mode == 'Never' or not mode:
                    row["Interrupt Audio"] = "Never"
                else:
                    row["Interrupt Audio"] = format_sec(q_set.get('holdAudioInterruptionPeriod'))

                vm_recip = str(rule.get('voicemail', {}).get('recipient', {}).get('id', ''))
                if vm_recip and vm_recip != 'None':
                    row["Voicemail Recipients"] = ext_id_to_num.get(vm_recip, vm_recip)

                if CQ_DEBUG:
                    print(f"[CQ_DEBUG] Audit greetings ext={row.get('Extension')}: "
                          f"{json.dumps(rule.get('greetings', []), default=str)[:2000]}", flush=True)

                for g in rule.get('greetings', []):
                    g_type = g.get('type')
                    g_preset = g.get('preset', {}) or {}
                    g_id = str(g_preset.get('id', ''))
                    g_embedded = str(g_preset.get('name', '')).strip()

                    is_music = g_type in ['ConnectingAudio', 'ConnectingMessage', 'HoldMusic']
                    if g_id:
                        type_key = 'ConnectingAudio' if g_type in ['ConnectingAudio', 'ConnectingMessage'] else g_type
                        mapped = preset_id_to_name.get(type_key, {}).get(g_id)
                        if mapped:
                            g_name = mapped
                        elif is_music:
                            # Music slots: only trust a genuine genre name; other embedded values
                            # (e.g. a "None"/"No" default marker) collapse to Default.
                            g_name = g_embedded.title() if g_embedded.lower() in AUDIO_GENRES else 'Default'
                        elif g_embedded:
                            g_name = g_embedded.title()
                        else:
                            g_name = 'Default'
                    elif 'custom' in g:
                        # Browser-downloadable service-web link (works with a logged-in portal
                        # session), built from the greeting id + owning extension id.
                        dl = _greeting_download_url(g.get('custom'))
                        g_name = f"Custom - {dl}" if dl else 'Custom'
                    else:
                        g_name = 'Default'

                    if g_type == 'Introductory': row["Greeting"] = g_name
                    elif g_type == 'ConnectingAudio':
                        # The real "Audio While Connecting" genre (e.g. Acoustic). ConnectingMessage
                        # is a separate slot and is intentionally ignored here — it was overwriting
                        # this value with unrelated content (e.g. "No").
                        row["Audio While Connecting"] = g_name
                    elif g_type == 'HoldMusic': row["Hold Music"] = g_name
                    elif g_type == 'InterruptPrompt':
                        if 'patience' in g_name.lower(): row["Interrupt Prompt"] = "Thank you for your patience"
                        elif 'volume' in g_name.lower(): row["Interrupt Prompt"] = "Higher than normal volume"
                        elif 'busy' in g_name.lower(): row["Interrupt Prompt"] = "Agents are currently busy"
                        elif 'important' in g_name.lower(): row["Interrupt Prompt"] = "Call is very important to us"
                        else: row["Interrupt Prompt"] = g_name
                    elif g_type == 'Voicemail': row["Voicemail Greeting"] = g_name

            succ, ah_rule = safe_api_call(f'/restapi/v1.0/account/~/extension/{qid}/answering-rule/after-hours-rule', token=token)
            if succ:
                row["After Hours Behavior"] = ah_rule.get('callHandlingAction')
                a_ext = _safe_get_ah_transfer_id(ah_rule.get('transfer'))
                if a_ext and a_ext != 'None':
                    row["After Hours Destination"] = ext_id_to_num.get(a_ext, a_ext)
                
                if not row.get("Voicemail Recipients"):
                    vm_recip_ah = str(ah_rule.get('voicemail', {}).get('recipient', {}).get('id', ''))
                    if vm_recip_ah and vm_recip_ah != 'None':
                        row["Voicemail Recipients"] = ext_id_to_num.get(vm_recip_ah, vm_recip_ah)

            succ, mgr_resp = safe_api_call(f'/restapi/v1.0/account/~/call-queues/{qid}/managers', token=token)
            if succ and mgr_resp.get('records'):
                mgrs = [ext_id_to_num.get(str(m.get('id', '')), '') for m in mgr_resp['records']]
                mgrs = [m for m in mgrs if m]
                if mgrs:
                    row["Queue Manager"] = ", ".join(mgrs)

            # Paginate (the endpoint caps a single page, so large queues lose the
            # overflow otherwise) and resolve each member the same robust way as
            # managers: prefer the record's extensionNumber, but fall back to mapping
            # its id -- some member records come back without extensionNumber, and
            # relying on it alone silently dropped those members from the export.
            succ, mem_records = fetch_directory(f'/restapi/v1.0/account/~/call-queues/{qid}/members', token)
            if succ and mem_records:
                mems = []
                for m in mem_records:
                    num = str(m.get('extensionNumber') or '').strip()
                    if not num:
                        num = ext_id_to_num.get(str(m.get('id', '')), '')
                    if num and num != 'None':
                        mems.append(num)
                if fixed_order_nums:
                    # Sequential queue: list agents in ring order, then any members not in it.
                    ordered = [n for n in fixed_order_nums if n in mems]
                    ordered += [n for n in mems if n not in fixed_order_nums]
                    mems = ordered
                row["Members (Ext)"] = ", ".join(mems)

            succ, notif = safe_api_call(f'/restapi/v1.0/account/~/extension/{qid}/notification-settings', token=token)
            if succ:
                vm_set = notif.get('voicemails', {})
                if vm_set.get('notifyByEmail'):
                    if vm_set.get('markAsRead'): row["Voicemail Notifications"] = "Notify Attach & Read"
                    elif vm_set.get('includeAttachment'): row["Voicemail Notifications"] = "Notify & Attach"
                    else: row["Voicemail Notifications"] = "Notify by Email"
                else:
                    row["Voicemail Notifications"] = "Off"
                
                # Basic mode keeps addresses at the top level; advanced mode keeps them under
                # the voicemails block (as emailAddresses or advancedEmailAddresses). Gather from
                # all of them so a queue on either setup still exports its addresses.
                if notif.get('advancedMode'):
                    emails = vm_set.get('emailAddresses') or vm_set.get('advancedEmailAddresses') or []
                else:
                    emails = notif.get('emailAddresses') or []
                if not emails:
                    emails = (vm_set.get('emailAddresses') or vm_set.get('advancedEmailAddresses')
                              or notif.get('emailAddresses') or [])
                emails = list(dict.fromkeys(emails))
                if emails: row["Voicemail Notifications Email"] = ", ".join(emails)

                # Per-type email toggles + voicemail-to-text, exported as On/Off so the
                # sheet round-trips back through the upload path.
                row["Voicemail to Text"] = "On" if vm_set.get('includeTranscription') else "Off"
                # Per-type On/Off + each type's own notification address (advanced mode keeps a
                # separate advancedEmailAddresses per type). Exported so the sheet round-trips.
                for _tcol, _ecol, _key in NOTIF_TYPE_COLS:
                    blk = notif.get(_key, {})
                    row[_tcol] = "On" if blk.get('notifyByEmail') else "Off"
                    _addrs = blk.get('advancedEmailAddresses') or blk.get('emailAddresses') or []
                    if _addrs:
                        row[_ecol] = ", ".join(dict.fromkeys(_addrs))

            rows.append(row)
            time.sleep(0.35) 

        df = pd.DataFrame(rows)
        for col in TEMPLATE_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        df = df[TEMPLATE_COLUMNS]

        output = build_config_workbook(df)
        audit_progress_store[task_id]['file_data'] = output.getvalue()
        audit_progress_store[task_id]['status'] = 'completed'
        audit_progress_store[task_id]['file_ready'] = True

    except Exception as e:
        audit_progress_store[task_id]['status'] = 'error'
        audit_progress_store[task_id]['error'] = str(e)

def get_val(row, key):
    clean_key = str(key).strip().lower()
    for k, v in row.items():
        if str(k).strip().lower() == clean_key:
            if pd.notna(v):
                val_str = str(v).strip()
                if val_str.lower() != 'nan' and val_str != '':
                    return val_str
    return None

def check_diff(changes_list, param_name, old_val, new_val):
    old_str = str(old_val).strip() if old_val is not None and str(old_val).strip() != '' else "None"
    new_str = str(new_val).strip() if new_val is not None and str(new_val).strip() != '' else "None"
    if old_str != new_str:
        changes_list.append({"parameter": param_name, "old": old_str, "new": new_str})
        return True
    return False

def update_cq_batch(records, token, is_preview=False, wipe_members=False, task_id=None, override_managers=False):
    total_records = len(records)
    yield {"type": "start", "total": total_records, "message": "Fetching Account Directories..."}
    
    succ, test_resp = safe_api_call('/restapi/v1.0/account/~', token=token)
    if not succ:
        yield {"type": "error", "message": f"Unauthorized. Token expired or invalid. Details: {format_api_error(test_resp)}"}
        return

    queue_map, ext_map, site_map, site_id_to_name = {}, {}, {}, {}
    tz_map, tz_id_to_name = {}, {}
    ext_id_to_num = {}
    
    succ, q_records = fetch_directory('/restapi/v1.0/account/~/call-queues', token)
    if succ:
        for q in q_records:
            if 'extensionNumber' in q: 
                queue_map[str(q['extensionNumber'])] = str(q['id'])
                ext_id_to_num[str(q['id'])] = str(q['extensionNumber'])
    else:
        yield {"type": "error", "message": "Failed to load queues directory."}
        return

    succ, e_records = fetch_directory('/restapi/v1.0/account/~/extension', token)
    if succ:
        for e in e_records:
            if 'extensionNumber' in e: 
                ext_map[str(e['extensionNumber'])] = str(e['id'])
                ext_id_to_num[str(e['id'])] = str(e['extensionNumber'])
    else:
        yield {"type": "error", "message": "Failed to load extensions directory."}
        return

    succ, s_records = fetch_directory('/restapi/v1.0/account/~/sites', token)
    if succ:
        for s in s_records:
            s_id = str(s['id'])
            s_name_dict = str(s.get('name')).lower().strip()
            site_map[s_name_dict] = s_id
            site_id_to_name[s_id] = str(s.get('name'))
            if s.get('code') == 'main-site' or s_name_dict == 'main site':
                site_map['main site'] = s_id
                site_map['company'] = s_id

    succ, tz_records = fetch_directory('/restapi/v1.0/dictionary/timezone', token)
    if succ:
        for tz in tz_records:
            tz_map[str(tz.get('name')).lower().strip()] = str(tz['id'])
            tz_map[str(tz.get('id'))] = str(tz['id'])
            tz_id_to_name[str(tz['id'])] = str(tz.get('name'))
    else:
        yield {"type": "error", "message": "Failed to load timezone dictionary."}
        return

    preset_dict = {'Introductory': {}, 'ConnectingAudio': {}, 'HoldMusic': {}, 'InterruptPrompt': {}, 'Voicemail': {}}
    preset_id_to_name = {'Introductory': {}, 'ConnectingAudio': {}, 'HoldMusic': {}, 'InterruptPrompt': {}, 'Voicemail': {}}
    
    for g_type in preset_dict.keys():
        succ1, dict1 = safe_api_call(f'/restapi/v1.0/dictionary/greeting?greetingType={g_type}&usageType=DepartmentExtensionAnsweringRule&perPage=1000', method='GET', token=token)
        succ2, dict2 = safe_api_call(f'/restapi/v1.0/dictionary/greeting?greetingType={g_type}&perPage=1000', method='GET', token=token)
        
        for dict_resp in [dict1, dict2]:
            if dict_resp and isinstance(dict_resp, dict) and 'records' in dict_resp:
                for rec in dict_resp['records']:
                    rec_type = rec.get('type')
                    if rec_type and rec_type != g_type:
                        continue 
                    k = rec.get('name', '').lower().strip()
                    v = str(rec.get('id', ''))
                    preset_dict[g_type][k] = v
                    preset_id_to_name[g_type][v] = str(rec.get('name', '')).title()

    def _resolve_ext(num):
        clean_num = str(num).split('.')[0].strip()
        if clean_num in ext_map: return ext_map[clean_num]
        return clean_num

    def _resolve_ext_id(num):
        """Resolve a member extension NUMBER to its RingCentral extension id, or
        None when no such extension exists. Members bulk-assign needs a real id --
        passing the raw number gets the whole call rejected with CMN-101, which
        drops every member in the request, so an unresolved number must be skipped
        rather than sent through. Falls back to a targeted lookup when the number
        isn't in the bulk directory (large accounts can page past the 1000-per-page
        fetch, leaving some extensions out of ext_map), caching any hit."""
        clean_num = str(num).split('.')[0].strip()
        if not clean_num:
            return None
        if clean_num in ext_map:
            return ext_map[clean_num]
        succ, resp = safe_api_call(f'/restapi/v1.0/account/~/extension?extensionNumber={clean_num}', method='GET', token=token)
        if succ and isinstance(resp, dict) and resp.get('records'):
            for rec in resp['records']:
                rid = str(rec.get('id', ''))
                if rid:
                    ext_map[clean_num] = rid
                    ext_id_to_num[rid] = clean_num
                    return rid
        return None

    def _dest_unresolved(num):
        """A transfer destination was supplied but doesn't map to a known extension.
        Left unresolved, RingCentral rejects the transfer and the action silently
        falls back to Voicemail, so callers can flag it as a clear error instead."""
        if num is None: return False
        clean_num = str(num).split('.')[0].strip()
        return bool(clean_num) and clean_num not in ext_map

    def _build_fixed_order_agents(queue_id, ordered_ext_ids):
        """Build the fixedOrderAgents array required when a queue switches to Sequential
        (FixedOrder). Order comes from the row's Members column when supplied, otherwise
        from the queue's current membership. The exact key RingCentral expects for ordering
        ('index') is not well documented; CQ_DEBUG surfaces the raw rejection if it's wrong."""
        ids = list(ordered_ext_ids or [])
        if not ids:
            succ, resp = safe_api_call(f'/restapi/v1.0/account/~/call-queues/{queue_id}/members', method='GET', token=token)
            if succ and isinstance(resp, dict):
                ids = [str(m.get('id', '')) for m in resp.get('records', []) if m.get('id')]
        return [{"extension": {"id": mid}, "index": idx} for idx, mid in enumerate(ids, start=1) if mid]

    for i, row in enumerate(records):
        # Cooperative stop (real apply only -- preview writes nothing). Queues
        # already synced cannot be recalled; remaining rows are simply skipped.
        if not is_preview and task_control.is_stopped(task_id):
            yield {"type": "cancelled", "current": i, "total": total_records,
                   "message": f"Stopped by user. {i} of {total_records} row(s) processed; the rest were skipped."}
            task_control.clear(task_id)
            return

        logs = []
        changes = []
        has_error = False

        ext_raw = get_val(row, 'Extension') or get_val(row, 'Extension Number')
        if not ext_raw: 
            yield {"type": "progress", "current": i + 1, "total": total_records, "result": {"ext": "N/A", "status": "info", "message": "Skipped row", "changes": []}, "is_preview": is_preview}
            continue
            
        ext_num = str(ext_raw).split('.')[0].strip()
        q_id = queue_map.get(ext_num)
        
        if not q_id:
            succ, resp = safe_api_call(f'/restapi/v1.0/account/~/extension?extensionNumber={ext_num}', method='GET', token=token)
            if succ and isinstance(resp, dict) and resp.get('records'):
                for rec in resp['records']:
                    if rec.get('type') == 'Department':
                        q_id = str(rec['id'])
                        queue_map[ext_num] = q_id
                        break
                        
        if not q_id:
            q_name = get_val(row, 'Queue Name')
            if not q_name:
                yield {"type": "progress", "current": i + 1, "total": total_records, "result": {"ext": ext_num, "status": "error", "message": "Call Queue not found. 'Queue Name' is required to create a new one.", "changes": []}, "is_preview": is_preview}
                continue
                
            if is_preview:
                q_id = f"mock_{ext_num}"
                queue_map[ext_num] = q_id
                ext_id_to_num[q_id] = ext_num
                changes.append({"parameter": "Queue", "old": "Missing", "new": "Will be created"})
                logs.append("Queue will be created")
            else:
                create_payload = {
                    "type": "Department",
                    "extensionNumber": ext_num,
                    "contact": { "firstName": q_name }
                }
                c_status = get_val(row, 'Status')
                if c_status: create_payload['status'] = c_status.capitalize()
                
                c_email = get_val(row, 'Queue Email')
                if c_email: create_payload['contact']['email'] = c_email
                    
                succ, c_resp = safe_api_call('/restapi/v1.0/account/~/extension', method='POST', json_payload=create_payload, token=token)
                if succ and isinstance(c_resp, dict) and c_resp.get('id'):
                    q_id = str(c_resp['id'])
                    queue_map[ext_num] = q_id
                    ext_map[ext_num] = q_id
                    ext_id_to_num[q_id] = ext_num
                    changes.append({"parameter": "Queue", "old": "Missing", "new": "Created"})
                    logs.append("Queue Created")
                    time.sleep(2.0)
                else:
                    yield {"type": "progress", "current": i + 1, "total": total_records, "result": {"ext": ext_num, "status": "error", "message": f"Failed to create Queue: {format_api_error(c_resp)}", "changes": changes}, "is_preview": is_preview}
                    continue

        # --- PRE-FETCH LEGACY ANSWERING RULE ---
        routing_fields = [
            'Ring Type', 'User Ring Time', 'Total Ring Time', 'Wrap Up Time', 
            'When Max Time is Reached', 'When Queue is Full', 'Callers In Queue', 
            'Interrupt Audio', 'Time Reached Destination', 'Queue Full Destination', 
            'Voicemail Recipients', 'Voicemail Greeting', 'Greeting', 'Audio While Connecting', 
            'Hold Music', 'Interrupt Prompt'
        ]
        
        orig_rule = {}
        if any(get_val(row, f) is not None for f in routing_fields):
            for _attempt in range(3):
                get_succ, rule = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/answering-rule/business-hours-rule', method='GET', token=token)
                if get_succ and isinstance(rule, dict):
                    orig_rule = copy.deepcopy(rule)
                    break
                time.sleep(2.0)

        # --- A. BASIC INFO UPDATE ---
        basic_fields = ['Queue Name', 'Status', 'Queue Email', 'Site', 'Timezone', 'Time Zone', 'Member Queue Status']
        if any(get_val(row, f) is not None for f in basic_fields):
            get_succ, old_basic = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}', method='GET', token=token)
            if get_succ and isinstance(old_basic, dict):
                basic_payload = {}
                b_needs_update = False

                # Rename a Call Queue the same way the proven extension_renamer module does
                # for non-User extensions: the display name lives in contact.firstName (with
                # lastName cleared so a stale last name can't concatenate), and the root `name`
                # is set too for consistency. The previous code set only the root `name`, which
                # RingCentral ignored for Department extensions, so the rename never applied.
                # Editing name and/or email both mutate the shared contact object, so build it
                # once from a copy of the current contact to avoid aliasing old_basic (which
                # would make check_diff compare a value against itself). Drop pronouncedName —
                # the renamer strips it because sending it back triggers CMN-101 errors.
                old_contact = {k: v for k, v in (old_basic.get('contact') or {}).items()
                               if k != 'pronouncedName'}

                val_qn = get_val(row, 'Queue Name')
                if val_qn is not None:
                    basic_payload.setdefault('contact', copy.deepcopy(old_contact))
                    b_needs_update |= check_diff(changes, 'Queue Name', old_basic.get('name'), val_qn)
                    basic_payload['contact']['firstName'] = val_qn
                    basic_payload['contact']['lastName'] = ""
                    basic_payload['name'] = val_qn

                val_st = get_val(row, 'Status')
                if val_st is not None:
                    basic_payload['status'] = val_st.capitalize()
                    b_needs_update |= check_diff(changes, 'Status', old_basic.get('status'), basic_payload['status'])

                val_qe = get_val(row, 'Queue Email')
                if val_qe is not None:
                    basic_payload.setdefault('contact', copy.deepcopy(old_contact))
                    b_needs_update |= check_diff(changes, 'Queue Email', old_contact.get('email'), val_qe)
                    basic_payload['contact']['email'] = val_qe

                val_site = get_val(row, 'Site')
                if val_site is not None:
                    s_name = val_site.lower()
                    if not site_map:
                        pass
                    else:
                        new_site_id = site_map.get(s_name)
                        if new_site_id: 
                            old_site_obj = old_basic.get('site', {})
                            old_site_id = str(old_site_obj.get('id', '')) if old_site_obj else ''
                            
                            if not old_site_id or old_site_id == 'None':
                                old_site_id = site_map.get('main site', 'main-site')

                            old_site_name = 'Main Site' if old_site_id in ('main-site', site_map.get('main site')) else site_id_to_name.get(old_site_id, old_site_id)
                            new_site_name = 'Main Site' if new_site_id in ('main-site', site_map.get('main site')) else site_id_to_name.get(new_site_id, new_site_id)
                            
                            if check_diff(changes, 'Site', old_site_name, new_site_name):
                                if new_site_id == 'main-site' or new_site_id == site_map.get('main site'):
                                    pass 
                                else:
                                    basic_payload['site'] = {'id': new_site_id}
                                    b_needs_update = True
                        else:
                            has_error = True
                            logs.append(f"Invalid Site: '{val_site}'")
                        
                tz_raw = get_val(row, 'Timezone') or get_val(row, 'Time Zone')
                if tz_raw is not None:
                    def find_tz_id(raw_tz):
                        raw_lower = str(raw_tz).lower().strip()
                        if raw_lower in tz_map: return tz_map[raw_lower]
                        # Look for city name
                        city = raw_lower.split('/')[-1].replace('_', ' ')
                        for tz in tz_records:
                            if city in str(tz.get('name', '')).lower() or city in str(tz.get('description', '')).lower():
                                return str(tz['id'])
                        # Fallback for common mismatches
                        if 'eastern' in raw_lower and 'us' in raw_lower:
                            for tz in tz_records:
                                if 'eastern' in str(tz.get('name', '')).lower() and 'us' in str(tz.get('description', '')).lower():
                                    return str(tz['id'])
                        if 'melbourne' in raw_lower or 'sydney' in raw_lower or 'canberra' in raw_lower:
                            for tz in tz_records:
                                if 'sydney' in str(tz.get('name', '')).lower():
                                    return str(tz['id'])
                        return None

                    new_tz_id = find_tz_id(tz_raw)
                    if new_tz_id:
                        if 'regionalSettings' not in basic_payload: basic_payload['regionalSettings'] = {}
                        basic_payload['regionalSettings']['timezone'] = {'id': new_tz_id}
                        
                        old_tz_id = str(old_basic.get('regionalSettings', {}).get('timezone', {}).get('id', ''))
                        old_tz_name = tz_id_to_name.get(old_tz_id, old_tz_id) if old_tz_id else "None"
                        new_tz_name = tz_id_to_name.get(new_tz_id, new_tz_id)
                        
                        b_needs_update |= check_diff(changes, 'Timezone', old_tz_name, new_tz_name)
                    else:
                        has_error = True; logs.append(f"Invalid Timezone: '{tz_raw}'")

                if b_needs_update:
                    if not is_preview:
                        put_succ, err = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}', method='PUT', json_payload=basic_payload, token=token)
                        attempt = 0
                        while not put_succ and 'extensionId' in str(err) and attempt < 3:
                            time.sleep(2.0)
                            attempt += 1
                            put_succ, err = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}', method='PUT', json_payload=basic_payload, token=token)
                        
                        if put_succ:
                            logs.append("Basic Info Updated")
                        else:
                            has_error = True
                            logs.append(f"Basic Error: {format_api_error(err)}")
                            _debug_dump(logs, 'Basic Info PUT failed', payload=basic_payload, err=err)
                    else:
                        logs.append("Basic Info Evaluated")
                        
        val_mqs = get_val(row, 'Member Queue Status')
        if val_mqs is not None:
            get_succ, old_cq = safe_api_call(f'/restapi/v1.0/account/~/call-queues/{q_id}', method='GET', token=token)
            if get_succ and isinstance(old_cq, dict):
                mem_status = val_mqs.lower()
                new_editable = True if 'allowed' in mem_status and 'not' not in mem_status else False
                
                old_editable = old_cq.get('editableMemberStatus')
                old_status_str = 'Allowed' if old_editable else 'Not Allowed'
                new_status_str = 'Allowed' if new_editable else 'Not Allowed'
                
                if check_diff(changes, 'Member Queue Status', old_status_str, new_status_str):
                    if not is_preview:
                        cq_payload = {"editableMemberStatus": new_editable}
                        s_succ, err = safe_api_call(f'/restapi/v1.0/account/~/call-queues/{q_id}', method='PUT', json_payload=cq_payload, token=token)
                        if s_succ: 
                            logs.append("Member Queue Status Updated")
                        else: 
                            has_error = True
                            logs.append(f"Member Status Error: {format_api_error(err)}")
                    else:
                        logs.append("Member Queue Status Evaluated")

        # --- B. BUSINESS HOURS UPDATE ---
        hours_str = get_val(row, 'Hours')
        if hours_str is not None:
            try:
                weekly_ranges = parse_intuitive_hours(hours_str)
                old_hours_str = "Unknown"
                get_succ, old_hours_resp = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/business-hours', method='GET', token=token)
                
                if get_succ and isinstance(old_hours_resp, dict):
                    old_ranges = old_hours_resp.get('schedule', {}).get('weeklyRanges', {})
                    if not old_ranges and 'schedule' in old_hours_resp: old_hours_str = "24/7"
                    else: old_hours_str = format_schedule(old_ranges)

                new_hours_str = "24/7" if weekly_ranges == "24/7" else format_schedule(weekly_ranges)
                
                if old_hours_str != new_hours_str:
                    changes.append({"parameter": "Business Hours", "old": old_hours_str, "new": new_hours_str})
                    if not is_preview:
                        payload = {"schedule": {}} if weekly_ranges == "24/7" else {"schedule": {"weeklyRanges": weekly_ranges}}
                        s_succ, err = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/business-hours', method='PUT', json_payload=payload, token=token)
                        if s_succ: 
                            logs.append("Hours Updated")
                        else: 
                            has_error = True
                            logs.append(f"Hours Error: {format_api_error(err)}")
                    else:
                        logs.append("Hours Evaluated")
            except Exception as e:
                has_error = True; logs.append(f"Hours Parse Error: {str(e)}")

        # --- C. ROUTING, TIMERS & LEGACY AUDIO ---
        wants_routing = any(get_val(row, f) is not None for f in routing_fields)
        if wants_routing and not orig_rule:
            # The pre-fetch above failed, so silently dropping Ring Type / timers would
            # look like a no-op to the user. Surface it instead.
            has_error = True
            logs.append("Routing Error: Could not load the queue's business-hours answering rule (Ring Type & timers skipped).")
        if wants_routing and orig_rule:
            rule = copy.deepcopy(orig_rule)
            q_set = rule.get('queue', {})
            r_needs_update = False
            
            rt = get_val(row, 'Ring Type')
            if rt is not None:
                rt_lower = rt.lower()
                if 'simultaneous' in rt_lower: q_set['transferMode'] = 'Simultaneous'
                elif 'sequential' in rt_lower or 'fixed' in rt_lower: q_set['transferMode'] = 'FixedOrder'
                elif 'rotating' in rt_lower or 'idle' in rt_lower: q_set['transferMode'] = 'Rotating'
            
            val_urt = get_val(row, 'User Ring Time')
            if val_urt is not None: 
                parsed = parse_time_to_seconds(val_urt)
                if parsed is not None: q_set['agentTimeout'] = parsed
                
            val_trt = get_val(row, 'Total Ring Time')
            if val_trt is not None:
                parsed = parse_time_to_seconds(val_trt)
                if parsed is not None: q_set['holdTime'] = parsed
                
            val_wut = get_val(row, 'Wrap Up Time')
            if val_wut is not None:
                parsed = parse_time_to_seconds(val_wut)
                if parsed is not None: q_set['wrapUpTime'] = parsed
                
            val_ciq = get_val(row, 'Callers In Queue')
            if val_ciq is not None:
                parsed = to_int(val_ciq)
                if parsed is not None: q_set['maxCallers'] = parsed
                
            val_ia = get_val(row, 'Interrupt Audio')
            if val_ia is not None:
                if val_ia.lower() == 'never':
                    q_set['holdAudioInterruptionMode'] = 'Never'
                    q_set.pop('holdAudioInterruptionPeriod', None)
                else:
                    parsed = parse_time_to_seconds(val_ia)
                    if parsed is not None: 
                        q_set['holdAudioInterruptionMode'] = 'Periodically'
                        q_set['holdAudioInterruptionPeriod'] = parsed

            val_wmtr = get_val(row, 'When Max Time is Reached')
            if val_wmtr is not None:
                if val_wmtr == 'TransferToExtension':
                    tr_dest = get_val(row, 'Time Reached Destination')
                    if _dest_unresolved(tr_dest):
                        has_error = True
                        logs.append(f"Max Time Dest '{tr_dest}' is not a known extension (transfer will fall back to Voicemail).")
                    dest_id = _resolve_ext(tr_dest)
                    q_set['holdTimeExpirationAction'] = val_wmtr
                    _set_queue_transfer(q_set, 'HoldTimeExpiration', dest_id)
                else:
                    q_set['holdTimeExpirationAction'] = val_wmtr
                    _set_queue_transfer(q_set, 'HoldTimeExpiration', None)

            val_wqf = get_val(row, 'When Queue is Full')
            if val_wqf is not None:
                if val_wqf == 'TransferToExtension':
                    qf_dest = get_val(row, 'Queue Full Destination')
                    if _dest_unresolved(qf_dest):
                        has_error = True
                        logs.append(f"Queue Full Dest '{qf_dest}' is not a known extension (transfer will fall back to Voicemail).")
                    dest_id = _resolve_ext(qf_dest)
                    q_set['maxCallersAction'] = val_wqf
                    _set_queue_transfer(q_set, 'MaxCallers', dest_id)
                else:
                    q_set['maxCallersAction'] = val_wqf
                    _set_queue_transfer(q_set, 'MaxCallers', None)
                    
            # Lean payload: Strip everything we aren't specifically allowed to touch from the deep copy
            for f in ['positionInQueue', 'callback', 'callers', 'calledNumbers']:
                q_set.pop(f, None)

            # RingCentral rejects a per-agent ring time (agentTimeout) longer than the total
            # wait time (holdTime) on Sequential/Rotating queues, which fails the whole queue
            # PUT. Clamp it down and tell the user rather than lose the entire update.
            if q_set.get('transferMode') != 'Simultaneous':
                eff_agent = to_int(q_set.get('agentTimeout'))
                eff_hold = to_int(q_set.get('holdTime'))
                if eff_agent is not None and eff_hold is not None and eff_agent > eff_hold:
                    q_set['agentTimeout'] = eff_hold
                    logs.append(f"User Ring Time ({format_sec(eff_agent)}) exceeded Total Ring Time ({format_sec(eff_hold)}); clamped to {format_sec(eff_hold)}.")

            # fixedOrderAgents is only valid when transferMode == 'FixedOrder'. If the queue is
            # currently Sequential and the upload switches Ring Type to Simultaneous/Rotating, a
            # stale agent list left in the payload makes RingCentral reject the whole queue update,
            # so Ring Type AND every ring timer silently fail together. Drop it unless we stay FixedOrder.
            if q_set.get('transferMode') != 'FixedOrder':
                q_set.pop('fixedOrderAgents', None)
            elif not q_set.get('fixedOrderAgents'):
                # Switching TO Sequential requires a non-empty agent order for the PUT to be
                # valid. Seed it from current membership just to satisfy that; the real ring
                # order from the sheet's Members column is applied in section F2 below, after
                # membership has been settled.
                fo_agents = _build_fixed_order_agents(q_id, [])
                if fo_agents:
                    q_set['fixedOrderAgents'] = fo_agents
                    _debug_dump(logs, 'Seeded fixedOrderAgents', payload=fo_agents)

            rule['queue'] = q_set
            
            old_q = orig_rule.get('queue', {})
            
            tm_map = {'FixedOrder': 'Sequential', 'Simultaneous': 'Simultaneous', 'Rotating': 'Rotating'}
            old_tm = tm_map.get(old_q.get('transferMode'), old_q.get('transferMode'))
            new_tm = tm_map.get(q_set.get('transferMode'), q_set.get('transferMode'))
            
            if rt is not None: r_needs_update |= check_diff(changes, 'Ring Type', old_tm, new_tm)
            if val_urt is not None: r_needs_update |= check_diff(changes, 'User Ring Time', format_sec(old_q.get('agentTimeout')), format_sec(q_set.get('agentTimeout')))
            if val_trt is not None: r_needs_update |= check_diff(changes, 'Total Ring Time', format_sec(old_q.get('holdTime')), format_sec(q_set.get('holdTime')))
            if val_wut is not None: r_needs_update |= check_diff(changes, 'Wrap Up Time', format_sec(old_q.get('wrapUpTime')), format_sec(q_set.get('wrapUpTime')))
            if val_ciq is not None: r_needs_update |= check_diff(changes, 'Max Callers', old_q.get('maxCallers'), q_set.get('maxCallers'))
            if val_wqf is not None: r_needs_update |= check_diff(changes, 'Max Callers Action', old_q.get('maxCallersAction'), q_set.get('maxCallersAction'))
            
            old_f_id = _safe_get_transfer_id(old_q.get('transfer'), 'MaxCallers') or 'None'
            new_f_id = _safe_get_transfer_id(q_set.get('transfer'), 'MaxCallers') or 'None'
            if get_val(row, 'Queue Full Destination') is not None:
                r_needs_update |= check_diff(changes, 'Queue Full Dest', ext_id_to_num.get(old_f_id, old_f_id), ext_id_to_num.get(new_f_id, new_f_id))

            if val_wmtr is not None: r_needs_update |= check_diff(changes, 'Max Time Action', old_q.get('holdTimeExpirationAction'), q_set.get('holdTimeExpirationAction'))
            
            old_t_id = _safe_get_transfer_id(old_q.get('transfer'), 'HoldTimeExpiration') or 'None'
            new_t_id = _safe_get_transfer_id(q_set.get('transfer'), 'HoldTimeExpiration') or 'None'
            if get_val(row, 'Time Reached Destination') is not None:
                r_needs_update |= check_diff(changes, 'Max Time Dest', ext_id_to_num.get(old_t_id, old_t_id), ext_id_to_num.get(new_t_id, new_t_id))

            old_ia_mode = old_q.get('holdAudioInterruptionMode')
            old_ia_str = "Never" if old_ia_mode == 'Never' or not old_ia_mode else format_sec(old_q.get('holdAudioInterruptionPeriod'))
                
            new_ia_mode = q_set.get('holdAudioInterruptionMode')
            new_ia_str = "Never" if new_ia_mode == 'Never' or not new_ia_mode else format_sec(q_set.get('holdAudioInterruptionPeriod'))
                
            if val_ia is not None:
                r_needs_update |= check_diff(changes, 'Interrupt Audio', old_ia_str, new_ia_str)

            # Purge toxic and incompatible presets before mapping
            safe_greetings = []
            for g in orig_rule.get('greetings', []):
                g_type = g.get('type')
                g_id = str(g.get('preset', {}).get('id', ''))
                if g_id in ['139008', '134401', '131847', '131843', '131853']: continue
                safe_greetings.append(g)
            
            rule['greetings'] = safe_greetings

            def apply_legacy_audio(col_name, slot_type, dict_type):
                nonlocal r_needs_update
                val = get_val(row, col_name)
                if val is None: return
                
                new_val = val.lower().strip()
                rule['greetings'] = [g for g in rule['greetings'] if g.get('type') != slot_type]

                if new_val.startswith('custom'):
                    # "Custom" or "Custom - <url>" (as the audit exports it) means keep the queue's
                    # existing custom audio. We can't upload audio from a URL here, so preserve the
                    # original greeting and record no change.
                    orig_g = next((g for g in orig_rule.get('greetings', []) if g.get('type') == slot_type), None)
                    if orig_g:
                        rule['greetings'].append(orig_g)
                    return
                elif new_val in ['off', 'none', 'disable', 'disabled']:
                    # "Off" is NOT the same as omitting the greeting. Omitting a greeting type
                    # makes RingCentral fall back to the Default greeting (still enabled), which
                    # is why Greeting=Off looked ignored in the portal. Off is represented by the
                    # greeting type's "None" preset, so set it explicitly.
                    none_id = preset_dict.get(dict_type, {}).get('none')
                    if none_id:
                        rule['greetings'].append({"type": slot_type, "preset": {"id": str(none_id)}})
                    else:
                        logs.append(f"{col_name}: no 'None' preset available to turn it Off; left unchanged.")
                        _debug_dump(logs, f'No None/Off preset in dictionary for {dict_type}')
                elif new_val == 'default':
                    pass  # Omitting the greeting type reverts it to the Default greeting.
                else:
                    matched_id = preset_dict.get(dict_type, {}).get(new_val)
                    if not matched_id:
                        for n, gid in preset_dict.get(dict_type, {}).items():
                            if new_val in n or n in new_val:
                                matched_id = gid
                                break
                    if not matched_id and dict_type == 'InterruptPrompt':
                        for n, gid in preset_dict['InterruptPrompt'].items():
                            if "patience" in new_val and "patience" in n: matched_id = gid
                            elif "volume" in new_val and "volume" in n: matched_id = gid
                            elif "busy" in new_val and "busy" in n: matched_id = gid
                            elif "important" in new_val and "important" in n: matched_id = gid
                            
                    if matched_id:
                        rule['greetings'].append({"type": slot_type, "preset": {"id": str(matched_id)}})
                    else:
                        orig_g = next((g for g in orig_rule.get('greetings', []) if g.get('type') == slot_type), None)
                        if orig_g:
                            old_id = str(orig_g.get('preset', {}).get('id', ''))
                            if old_id not in ['139008', '134401', '131847', '131843', '131853']:
                                rule['greetings'].append(orig_g)
                
                old_val_name = get_old_greeting_name(orig_rule, slot_type)
                new_val_str = val.title() if new_val not in ['off', 'none', 'disable', 'disabled', 'default'] else 'Default'
                if new_val in ['off', 'none', 'disable', 'disabled']: new_val_str = 'Off'
                r_needs_update |= check_diff(changes, col_name, old_val_name, new_val_str)

            apply_legacy_audio('Greeting', 'Introductory', 'Introductory')
            apply_legacy_audio('Hold Music', 'HoldMusic', 'HoldMusic')
            apply_legacy_audio('Interrupt Prompt', 'InterruptPrompt', 'InterruptPrompt')
            apply_legacy_audio('Voicemail Greeting', 'Voicemail', 'Voicemail')
            apply_legacy_audio('Audio While Connecting', 'ConnectingAudio', 'ConnectingAudio')
            
            vm_recip_raw = get_val(row, 'Voicemail Recipients')
            if vm_recip_raw is not None:
                vm_ext_id = _resolve_ext(vm_recip_raw)
                if vm_ext_id:
                    if 'voicemail' not in rule: rule['voicemail'] = {}
                    rule['voicemail']['recipient'] = {'id': vm_ext_id}
                    old_vm = str(orig_rule.get('voicemail', {}).get('recipient', {}).get('id', 'None'))
                    r_needs_update |= check_diff(changes, 'VM Recipient', ext_id_to_num.get(old_vm, old_vm), ext_id_to_num.get(vm_ext_id, vm_ext_id))
            
            for field in _READ_ONLY: rule.pop(field, None)
            for field in _READ_ONLY: 
                if 'queue' in rule: rule['queue'].pop(field, None)
            
            if r_needs_update and not is_preview:
                put_succ, err = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/answering-rule/business-hours-rule', method='PUT', json_payload=rule, token=token)
                
                if not put_succ and ('transfer' in str(err) or 'transfer.extension.id' in str(err)):
                    rule['queue'].pop('transfer', None)
                    if rule['queue'].get('maxCallersAction') == 'TransferToExtension': rule['queue']['maxCallersAction'] = 'Voicemail'
                    if rule['queue'].get('holdTimeExpirationAction') == 'TransferToExtension': rule['queue']['holdTimeExpirationAction'] = 'Voicemail'
                    
                    put_succ2, err2 = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/answering-rule/business-hours-rule', method='PUT', json_payload=rule, token=token)
                    if put_succ2:
                        logs.append("Routing Updated (Invalid transfers stripped & reverted to Voicemail)")
                    else:
                        has_error = True
                        logs.append(f"Routing Error: {format_api_error(err2)}")
                        _debug_dump(logs, 'Routing PUT (retry) failed', payload=rule.get('queue'), err=err2)

                elif put_succ:
                    logs.append("Routing Updated")
                else:
                    has_error = True
                    logs.append(f"Routing Error: {format_api_error(err)}")
                    _debug_dump(logs, 'Routing PUT failed', payload=rule.get('queue'), err=err)

        # --- E. AFTER HOURS RULE ---
        ah_fields = ['After Hours Behavior', 'After Hours Destination']
        if any(get_val(row, f) is not None for f in ah_fields):
            orig_ah = {}
            for _attempt in range(3):
                get_succ, ah_rule_resp = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/answering-rule/after-hours-rule', method='GET', token=token)
                if get_succ and isinstance(ah_rule_resp, dict):
                    orig_ah = copy.deepcopy(ah_rule_resp)
                    break
                time.sleep(2.0)
            
            if orig_ah:
                ah_rule = copy.deepcopy(orig_ah)
                a_needs_update = False
                
                val_ahb = get_val(row, 'After Hours Behavior')
                if val_ahb is not None:
                    # RingCentral's answering-rule enum uses 'TakeMessagesOnly' for
                    # voicemail. Accept the human-friendly 'Voicemail' too so a value
                    # typed into the sheet doesn't get rejected with CMN-101.
                    if val_ahb.strip().lower() in ('voicemail', 'takemessagesonly'):
                        val_ahb = 'TakeMessagesOnly'
                    if val_ahb == 'TransferToExtension':
                        ah_dest = get_val(row, 'After Hours Destination')
                        if _dest_unresolved(ah_dest):
                            has_error = True
                            logs.append(f"After Hours Dest '{ah_dest}' is not a known extension (transfer will fall back to Voicemail).")
                        dest_id = _resolve_ext(ah_dest)
                        ah_rule['callHandlingAction'] = val_ahb
                        # transfer is an object, not a list (see bulk_hours _build_rule_api_body).
                        if dest_id: ah_rule['transfer'] = {'extension': {'id': dest_id}}
                    else:
                        ah_rule['callHandlingAction'] = val_ahb
                        ah_rule.pop('transfer', None)
                
                if val_ahb is not None:
                    a_needs_update |= check_diff(changes, 'After Hours Behavior', orig_ah.get('callHandlingAction'), ah_rule.get('callHandlingAction'))
                
                old_a_id = _safe_get_ah_transfer_id(orig_ah.get('transfer')) or 'None'
                new_a_id = _safe_get_ah_transfer_id(ah_rule.get('transfer')) or 'None'
                if get_val(row, 'After Hours Destination') is not None:
                    a_needs_update |= check_diff(changes, 'After Hours Dest', ext_id_to_num.get(old_a_id, old_a_id), ext_id_to_num.get(new_a_id, new_a_id))
                
                for field in _READ_ONLY: ah_rule.pop(field, None)
                ah_rule.pop('greetings', None); ah_rule.pop('callers', None); ah_rule.pop('calledNumbers', None)
                
                if a_needs_update and not is_preview:
                    put_succ, err = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/answering-rule/after-hours-rule', method='PUT', json_payload=ah_rule, token=token)
                    
                    if not put_succ and ('transfer' in str(err) or 'transfer.extension.id' in str(err)):
                        ah_rule.pop('transfer', None)
                        # 'TakeMessagesOnly' is the API's voicemail value; 'Voicemail' is rejected (CMN-101).
                        ah_rule['callHandlingAction'] = 'TakeMessagesOnly'
                        put_succ2, err2 = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/answering-rule/after-hours-rule', method='PUT', json_payload=ah_rule, token=token)
                        if put_succ2: logs.append("After Hours Updated (Invalid transfer stripped & reverted to Voicemail)")
                        else:
                            has_error = True; logs.append(f"After Hours Error: {format_api_error(err2)}")
                            _debug_dump(logs, 'After Hours PUT (retry) failed', payload=ah_rule, err=err2)

                    elif put_succ:
                        logs.append("After Hours Updated")
                    else:
                        has_error = True
                        logs.append(f"After Hours Error: {format_api_error(err)}")
                        _debug_dump(logs, 'After Hours PUT failed', payload=ah_rule, err=err)

        # --- F. MEMBERS ---
        # A blank Members column is always left untouched. When it has values the sheet
        # members are added; the wipe_members toggle additionally removes any current
        # member not listed, making the queue's membership exactly the sheet list.
        members_changed = False
        val_mems = get_val(row, 'Members (Ext)')
        if val_mems is not None:
            mem_exts = [e.strip() for e in val_mems.split(',') if e.strip()]
            if mem_exts:
                old_mem_list = []
                old_id_by_num = {}
                get_succ, old_mems_resp = safe_api_call(f'/restapi/v1.0/account/~/call-queues/{q_id}/members', method='GET', token=token)
                if get_succ and isinstance(old_mems_resp, dict):
                    for m in old_mems_resp.get('records', []):
                        num = str(m.get('extensionNumber'))
                        if num and num != 'None':
                            old_mem_list.append(num)
                            if m.get('id'): old_id_by_num[num] = str(m.get('id'))

                old_set = set(old_mem_list)
                new_set = set(mem_exts)

                if wipe_members:
                    # Membership becomes exactly the sheet list.
                    target_list = mem_exts
                    remove_nums = [n for n in old_mem_list if n not in new_set]
                else:
                    # Additive: keep everyone already assigned, append the new ones.
                    target_list = old_mem_list + [n for n in mem_exts if n not in old_set]
                    remove_nums = []

                add_nums = [n for n in mem_exts if n not in old_set]
                target_set = set(target_list)

                if target_set != old_set:
                    old_mems_str = ", ".join(old_mem_list) if old_mem_list else "None"
                    new_mems_str = ", ".join(target_list) if target_list else "None"
                    changes.append({"parameter": "Queue Members", "old": old_mems_str, "new": new_mems_str})
                    if not is_preview:
                        added_ids, unresolved_add = [], []
                        for n in add_nums:
                            rid = _resolve_ext_id(n)
                            if rid:
                                added_ids.append(rid)
                            else:
                                unresolved_add.append(str(n).split('.')[0].strip())
                        removed_ids = [old_id_by_num[n] for n in remove_nums if n in old_id_by_num]

                        if unresolved_add:
                            # A single member number that isn't a real extension makes
                            # RingCentral reject the entire bulk-assign (CMN-101), which
                            # silently drops every other member too. Skip the unknown
                            # numbers, assign the ones that resolve, and flag the row so
                            # the bad extension is noticed instead of the queue coming up
                            # with no members.
                            has_error = True
                            logs.append(f"Members Warning: the following member(s) were removed as they do not exist: {', '.join(unresolved_add)}")

                        if added_ids or removed_ids:
                            # RingCentral's bulk-assign only applies reliably when BOTH keys are
                            # present. Sending addedExtensionIds alone (the additive case, where
                            # nothing is removed) left the queue unchanged, so always include both,
                            # empty array when unused.
                            mem_payload = {
                                "addedExtensionIds": added_ids,
                                "removedExtensionIds": removed_ids,
                            }
                            _debug_dump(logs, f'Members bulk-assign (wipe={wipe_members})', payload=mem_payload)
                            s_succ, err = safe_api_call(f'/restapi/v1.0/account/~/call-queues/{q_id}/bulk-assign', method='POST', json_payload=mem_payload, token=token)
                            if s_succ:
                                logs.append("Members Updated")
                                members_changed = True
                            else:
                                has_error = True
                                logs.append(f"Members Error: {format_api_error(err)}")
                                _debug_dump(logs, 'Members bulk-assign failed', payload=mem_payload, err=err)

        # --- F2. FIXED-ORDER RING SEQUENCE ---
        # For Sequential (FixedOrder) queues the agent order decides who rings first, so the
        # sheet's Members order is significant. Apply it here, after section F has settled
        # membership, so every listed agent is already a member when the sequence is set.
        mem_order_raw = get_val(row, 'Members (Ext)')
        if mem_order_raw:
            ordered_ids = [rid for rid in (_resolve_ext_id(e.strip()) for e in mem_order_raw.split(',') if e.strip()) if rid]
            if members_changed and not is_preview:
                time.sleep(2.0)  # let a just-changed membership settle before referencing it
            get_succ, fo_rule = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/answering-rule/business-hours-rule', method='GET', token=token)
            if get_succ and isinstance(fo_rule, dict) and fo_rule.get('queue', {}).get('transferMode') == 'FixedOrder' and ordered_ids:
                cur_ids = [str(a.get('extension', {}).get('id', '')) for a in fo_rule['queue'].get('fixedOrderAgents', [])]
                des_ids = [str(m) for m in ordered_ids]
                if cur_ids != des_ids:
                    old_nums = ", ".join(ext_id_to_num.get(i, i) for i in cur_ids) or "None"
                    new_nums = ", ".join(ext_id_to_num.get(i, i) for i in des_ids)
                    changes.append({"parameter": "Ring Order", "old": old_nums, "new": new_nums})
                    if not is_preview:
                        fo_rule['queue']['fixedOrderAgents'] = [{"extension": {"id": m}, "index": idx} for idx, m in enumerate(des_ids, start=1)]
                        for f in _READ_ONLY: fo_rule.pop(f, None)
                        for f in _READ_ONLY:
                            if 'queue' in fo_rule: fo_rule['queue'].pop(f, None)
                        put_succ, err = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/answering-rule/business-hours-rule', method='PUT', json_payload=fo_rule, token=token)
                        if put_succ:
                            logs.append("Ring Order Updated")
                        else:
                            has_error = True
                            logs.append(f"Ring Order Error: {format_api_error(err)}")
                            _debug_dump(logs, 'Fixed order PUT failed', payload=fo_rule.get('queue'), err=err)

        # --- G. VOICEMAIL & EMAIL NOTIFICATIONS ---
        # Per-type email toggles (missed calls, faxes, texts) plus the voicemail-to-text
        # transcription switch, alongside the original voicemail-notification handling.
        # Any one of these columns triggers a notification-settings PUT; untouched
        # columns are left exactly as the queue already has them.
        notif_toggle_cols = [(tcol, key) for tcol, _ecol, key in NOTIF_TYPE_COLS]
        vm_fields = (['Voicemail Notifications', 'Voicemail Notifications Email', 'Queue Email',
                      'Voicemail to Text'] + [c for c, _ in notif_toggle_cols]
                     + [ecol for _tcol, ecol, _key in NOTIF_TYPE_COLS])
        if any(get_val(row, f) is not None for f in vm_fields):
            get_succ, notif = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/notification-settings', method='GET', token=token)
            if get_succ and isinstance(notif, dict):
                orig_notif = copy.deepcopy(notif)
                vm_set = notif.get('voicemails', {})
                v_needs_update = False

                # Manager-recipient guard. There are two notification setups: (1) a real user
                # Manager is selected -- RingCentral delivers by manager reference, so email is on
                # but NO editable recipient address is stored; (2) a Specified Email (including the
                # dummy address Pro Serv provisions "as" the manager via API) -- a real address is
                # stored. Case (2) has a stored address and is updated as a normal specified-email
                # queue. Case (1) can't be edited normally (no address, and advanced mode rejects
                # manager-based recipients), so by default we skip it -- UNLESS the operator ticked
                # "Convert manager queues" AND the sheet supplies a specified address: then we untick
                # the manager, switch to advanced mode, and set the sheet's specified emails.
                _ovm = orig_notif.get('voicemails', {})
                _existing_addr = (_ovm.get('emailAddresses') or _ovm.get('advancedEmailAddresses')
                                  or orig_notif.get('emailAddresses') or [])
                _sheet_emails = ([get_val(row, 'Voicemail Notifications Email'), get_val(row, 'Queue Email')]
                                 + [get_val(row, ecol) for _tcol, ecol, _key in NOTIF_TYPE_COLS])
                _has_sheet_email = any(e for e in _sheet_emails)
                _is_manager = bool(_ovm.get('notifyByEmail')) and not _existing_addr
                manager_override = _is_manager and override_managers and _has_sheet_email
                vm_skip = _is_manager and not manager_override
                if vm_skip:
                    _msg = ("Manager recipient, no editable address" if not override_managers
                            else "Manager recipient -- tick 'Convert' and supply an email to override")
                    changes.append({"parameter": "VM Notifications", "old": "Manager recipient",
                                    "new": "Skipped", "skipped": True})
                    logs.append(f"Skipped VM notifications ({_msg})")
                if manager_override:
                    changes.append({"parameter": "VM Notifications", "old": "Manager recipient",
                                    "new": "Converted to Specified Emails"})
                    logs.append("Converted manager queue to Specified Emails (advanced mode)")

                val_vn = get_val(row, 'Voicemail Notifications')
                # "Include attachment" and "Mark as read" are advanced-mode-only settings in
                # RingCentral -- basic mode exposes only an on/off "By email" tick, so those flags
                # are silently ignored on a basic-mode queue. When the requested style needs
                # either, we must switch the queue to advanced mode for them to actually apply.
                want_advanced = False
                if val_vn is not None:
                    vm_val = val_vn.lower()
                    if vm_val in ['off', 'false', 'no']:
                        vm_set['notifyByEmail'] = False; vm_set['includeAttachment'] = False; vm_set['markAsRead'] = False
                    elif 'read' in vm_val:
                        vm_set['notifyByEmail'] = True; vm_set['includeAttachment'] = True; vm_set['markAsRead'] = True
                        want_advanced = True
                    elif 'attach' in vm_val:
                        vm_set['notifyByEmail'] = True; vm_set['includeAttachment'] = True; vm_set['markAsRead'] = False
                        want_advanced = True
                    else:
                        vm_set['notifyByEmail'] = True; vm_set['includeAttachment'] = False; vm_set['markAsRead'] = False

                # Distinct per-type notification addresses, and converting a manager queue, both
                # require advanced mode (basic mode has only one shared address / manager tick).
                _pertype_email_given = any(get_val(row, ecol) for _tcol, ecol, _key in NOTIF_TYPE_COLS)
                if manager_override or _pertype_email_given:
                    want_advanced = True

                new_emails = []
                val_vne = get_val(row, 'Voicemail Notifications Email')
                if val_vne is not None:
                    new_emails = [e.strip() for e in val_vne.split(',') if e.strip()]

                if vm_set.get('notifyByEmail'):
                    if not new_emails:
                        fallback = get_val(row, 'Queue Email')
                        if fallback:
                            new_emails = [e.strip() for e in fallback.split(',') if e.strip()]
                        else:
                            # Reuse the queue's existing notification address so changing the
                            # notification style (e.g. -> "Notify & Attach") with a blank email
                            # column doesn't wipe email notifications. Gather from every place the
                            # export path reads: advanced-mode queues keep addresses under
                            # voicemails.emailAddresses / advancedEmailAddresses, basic mode at the
                            # top level. Only disable email notification when no address exists at
                            # all (RingCentral rejects notifyByEmail=True with no address).
                            if notif.get('advancedMode'):
                                existing_emails = vm_set.get('emailAddresses') or vm_set.get('advancedEmailAddresses') or []
                            else:
                                existing_emails = notif.get('emailAddresses') or []
                            if not existing_emails:
                                existing_emails = (vm_set.get('emailAddresses') or vm_set.get('advancedEmailAddresses')
                                                   or notif.get('emailAddresses') or [])
                            if existing_emails:
                                new_emails = list(existing_emails)
                            else:
                                vm_set['notifyByEmail'] = False
                                vm_set['includeAttachment'] = False
                                vm_set['markAsRead'] = False

                if not vm_set.get('notifyByEmail'):
                    vm_set['includeAttachment'] = False
                    vm_set['markAsRead'] = False

                new_notif = copy.deepcopy(orig_notif)
                for field in _READ_ONLY: new_notif.pop(field, None)

                if 'voicemails' not in new_notif:
                    new_notif['voicemails'] = {}

                new_notif['voicemails']['notifyByEmail'] = vm_set.get('notifyByEmail', False)

                # Switch to advanced mode when the requested style needs attach/read, when
                # per-type addresses are supplied, or when converting a manager queue. These
                # features only exist under advanced mode; a basic-mode queue ignores them.
                if want_advanced and (vm_set.get('notifyByEmail') or manager_override or _pertype_email_given):
                    new_notif['advancedMode'] = True
                    # In advanced mode recipients live per-notification-type. The shared
                    # top-level emailAddresses is a basic-mode field and RingCentral rejects it
                    # here with CMN-101, so drop it -- per-type advancedEmailAddresses are used
                    # instead.
                    new_notif.pop('emailAddresses', None)
                    if manager_override:
                        # Untick the manager: turn off manager delivery at the top level and on
                        # every type, so notifications go only to the sheet's specified emails.
                        new_notif['includeManagers'] = False
                    # else: leave top-level includeManagers as-is to preserve manager delivery of
                    # any other notification types that relied on it.

                # Only rewrite the shared address list when the sheet supplied one (or a
                # fallback filled it). A blank email column must never wipe the queue's
                # existing addresses -- which matters now that a fax/text toggle alone can
                # trigger this PUT.
                if new_emails:
                    if new_notif.get('advancedMode'):
                        # Advanced mode stores the voicemail recipient under the voicemails block
                        # as advancedEmailAddresses (a list of address strings). The basic-mode
                        # emailAddresses fields -- top-level and per-type -- are invalid here
                        # (CMN-101), so use advancedEmailAddresses only and drop both emailAddresses.
                        new_notif['voicemails'].pop('emailAddresses', None)
                        new_notif['voicemails']['advancedEmailAddresses'] = new_emails
                        new_notif.pop('emailAddresses', None)
                    else:
                        new_notif['emailAddresses'] = new_emails

                # Voicemail-to-text. An explicit column value wins; when it's blank we keep
                # the long-standing default of transcription-on for rows that (re)set the
                # voicemail notification, but leave transcription untouched for rows that
                # only flip another notification type.
                vm_to_text = _parse_toggle(get_val(row, 'Voicemail to Text'))
                if vm_set.get('notifyByEmail'):
                    new_notif['voicemails']['includeAttachment'] = vm_set.get('includeAttachment', False)
                    new_notif['voicemails']['markAsRead'] = vm_set.get('markAsRead', False)
                    if vm_to_text is not None:
                        want_trans = vm_to_text
                    elif val_vn is not None:
                        want_trans = True
                    else:
                        want_trans = bool(orig_notif.get('voicemails', {}).get('includeTranscription', False))
                    new_notif['voicemails']['includeTranscription'] = want_trans
                else:
                    new_notif['voicemails'].pop('includeAttachment', None)
                    new_notif['voicemails'].pop('markAsRead', None)
                    new_notif['voicemails'].pop('includeTranscription', None)
                    want_trans = False

                old_email_on = str(orig_notif.get('voicemails', {}).get('notifyByEmail'))
                new_email_on = str(vm_set.get('notifyByEmail'))
                if val_vn is not None and not vm_skip:
                    v_needs_update |= check_diff(changes, 'VM Email On', old_email_on, new_email_on)
                    v_needs_update |= check_diff(changes, 'VM Attach', str(orig_notif.get('voicemails', {}).get('includeAttachment')), str(vm_set.get('includeAttachment')))
                    # markAsRead is the only difference between "Notify & Attach" and
                    # "Notify Attach & Read"; diff it separately so switching between the two
                    # is both visible in the preview and able to trigger the PUT on its own
                    # (otherwise a read-only change with no other diff was silently dropped).
                    v_needs_update |= check_diff(changes, 'VM Mark Read', str(orig_notif.get('voicemails', {}).get('markAsRead')), str(vm_set.get('markAsRead')))

                if (val_vn is not None or vm_to_text is not None) and not vm_skip:
                    old_trans = str(orig_notif.get('voicemails', {}).get('includeTranscription', False))
                    v_needs_update |= check_diff(changes, 'VM Transcription', old_trans, str(want_trans))

                # Per-type email notification toggles (missed calls, faxes, texts). Each
                # column flips notifyByEmail for that notification type only. Skipped for
                # Manager-recipient queues (see the manager-recipient guard above).
                for col_name, key in (notif_toggle_cols if not vm_skip else []):
                    tog = _parse_toggle(get_val(row, col_name))
                    if tog is None:
                        continue
                    if not isinstance(new_notif.get(key), dict):
                        new_notif[key] = {}
                    old_on = str(orig_notif.get(key, {}).get('notifyByEmail'))
                    new_notif[key]['notifyByEmail'] = tog
                    v_needs_update |= check_diff(changes, col_name, old_on, str(tog))

                old_emails = orig_notif.get('voicemails', {}).get('emailAddresses', []) if orig_notif.get('advancedMode') else orig_notif.get('emailAddresses', [])
                if val_vne is not None or get_val(row, 'Queue Email') is not None:
                    if set(old_emails) != set(new_emails):
                        v_needs_update |= check_diff(changes, 'VM Emails', ", ".join(old_emails), ", ".join(new_emails))

                # Advanced mode requires EVERY email-enabled notification type to carry a
                # non-empty advancedEmailAddresses (EXT-455). For each type that is ON, set its
                # address in priority order: (1) that type's own sheet column (e.g. "Text
                # Notifications Email"), (2) the queue's cascade address -- the sheet's Voicemail
                # Notifications Email when supplied, else the queue's existing notification address.
                # Each type's on/off state is left as-is unless the per-type toggle changed it.
                if new_notif.get('advancedMode'):
                    cascade_addr = (new_emails or orig_notif.get('emailAddresses') or [])
                    if not cascade_addr:
                        # No voicemail/existing address (e.g. converting a manager queue supplying
                        # only per-type emails) -- fall back to any address the sheet provided so
                        # every enabled type still validates in advanced mode.
                        for _c in (['Voicemail Notifications Email', 'Queue Email']
                                   + [ecol for _tcol, ecol, _key in NOTIF_TYPE_COLS]):
                            _cv = get_val(row, _c)
                            if _cv:
                                cascade_addr = [e.strip() for e in _cv.split(',') if e.strip()]
                                break
                    sheet_addr_given = (val_vne is not None and bool(new_emails)) or manager_override
                    # Map RingCentral keys -> that type's own sheet email column (voicemail's is
                    # the VM email, already reflected in cascade_addr).
                    pertype_addr = {}
                    for _tcol, _ecol, _key in NOTIF_TYPE_COLS:
                        _v = get_val(row, _ecol)
                        if _v:
                            pertype_addr[_key] = [e.strip() for e in _v.split(',') if e.strip()]
                    for _t in ('voicemails', 'inboundFaxes', 'outboundFaxes', 'inboundTexts', 'missedCalls', 'callNotes'):
                        blk = new_notif.get(_t)
                        if not (isinstance(blk, dict) and blk.get('notifyByEmail')):
                            continue
                        if manager_override:
                            blk['includeManagers'] = False
                        if _t in pertype_addr:
                            blk.pop('emailAddresses', None)
                            blk['advancedEmailAddresses'] = list(pertype_addr[_t])
                        elif sheet_addr_given or not blk.get('advancedEmailAddresses'):
                            blk.pop('emailAddresses', None)
                            blk['advancedEmailAddresses'] = list(cascade_addr)

                    # Surface the mode switch and any per-type address change so they're visible
                    # in the preview AND can trigger the PUT on their own (e.g. a per-type email
                    # supplied for a type that is already on produces no other diff).
                    if not vm_skip:
                        v_needs_update |= check_diff(changes, 'VM Advanced Mode',
                                                     str(orig_notif.get('advancedMode')), str(new_notif.get('advancedMode')))
                        for _tcol, _ecol, _key in NOTIF_TYPE_COLS:
                            if get_val(row, _ecol):
                                _old_a = ", ".join(orig_notif.get(_key, {}).get('advancedEmailAddresses')
                                                   or orig_notif.get(_key, {}).get('emailAddresses') or [])
                                _new_a = ", ".join(new_notif.get(_key, {}).get('advancedEmailAddresses') or [])
                                v_needs_update |= check_diff(changes, f"{_tcol} Email", _old_a, _new_a)

                if (v_needs_update or manager_override) and not vm_skip and not is_preview:
                    put_succ, err = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/notification-settings', method='PUT', json_payload=new_notif, token=token)
                    
                    if not put_succ and ('includeAttachment' in str(err) or 'markAsRead' in str(err) or 'includeTranscription' in str(err)):
                        new_notif['voicemails'].pop('includeAttachment', None)
                        new_notif['voicemails'].pop('markAsRead', None)
                        new_notif['voicemails'].pop('includeTranscription', None)
                        
                        put_succ2, err2 = safe_api_call(f'/restapi/v1.0/account/~/extension/{q_id}/notification-settings', method='PUT', json_payload=new_notif, token=token)
                        if put_succ2:
                            logs.append("Notifications Updated (Attachments/Transcription popped due to account limits)")
                        else:
                            has_error = True
                            logs.append(f"Notifications Error: {format_api_error(err2)}")
                            _debug_dump(logs, 'Notifications PUT (retry) failed', payload=new_notif, err=err2)
                    elif put_succ:
                        logs.append("Notifications Updated")
                    else:
                        has_error = True
                        logs.append(f"Notifications Error: {format_api_error(err)}")
                        _debug_dump(logs, 'Notifications PUT failed', payload=new_notif, err=err)

        if not logs and not changes: 
            res_dict = {"ext": ext_num, "status": "info", "message": "No valid changes found in row.", "changes": changes}
        elif has_error: 
            res_dict = {"ext": ext_num, "status": "error", "message": " | ".join(logs) or "Unknown Error", "changes": changes}
        else: 
            res_dict = {"ext": ext_num, "status": "success", "message": "Evaluated successfully." if is_preview else "Changes synced.", "changes": changes}
            
        yield {"type": "progress", "current": i + 1, "total": total_records, "result": res_dict, "is_preview": is_preview}
        time.sleep(1.5)

    task_control.clear(task_id)
    yield {"type": "done", "is_preview": is_preview}
