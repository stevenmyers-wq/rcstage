import io
import pandas as pd
from datetime import datetime
from flask import Blueprint, jsonify, send_file
from openpyxl.styles import PatternFill
from webapp.auth_utils import require_rc_token, get_rc_access_token
from webapp.usage_tracking import track_usage
from . import utils

user_timezone_audit_bp = Blueprint(
    'user_timezone_audit_bp', __name__, url_prefix='/api/user_timezone_audit'
)


@user_timezone_audit_bp.route('/export', methods=['POST'])
@require_rc_token
@track_usage('User Timezone Audit - Export')
def export_timezone_audit():
    token = get_rc_access_token()
    if not token:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        data = utils.generate_timezone_audit(token)
        df = pd.DataFrame(data)

        # Surface mismatches first, then unknowns, then matches; keep sites
        # grouped within each bucket.
        if not df.empty and "Match (Yes/No)" in df.columns:
            order = {"No": 0, "Unknown": 1, "Yes": 2}
            df["_sort"] = df["Match (Yes/No)"].map(lambda v: order.get(v, 3))
            df.sort_values(by=["_sort", "Site", "Ext Number"], inplace=True)
            df.drop(columns=["_sort"], inplace=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='User Timezone Audit')
            worksheet = writer.sheets['User Timezone Audit']

            # Auto-adjust column widths.
            for column in worksheet.columns:
                length = max(len(str(cell.value) or "") for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = min(length + 5, 50)

            # Highlight mismatched rows in red for quick scanning.
            headers = [c.value for c in worksheet[1]]
            if "Match (Yes/No)" in headers:
                match_col = headers.index("Match (Yes/No)") + 1
                red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for row in range(2, worksheet.max_row + 1):
                    if worksheet.cell(row=row, column=match_col).value == "No":
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row, column=col).fill = red

        output.seek(0)
        filename = f"User_Timezone_Audit_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
