import io
import time
import pandas as pd
from openpyxl.styles import PatternFill, Font
from webapp.rc_api import rc_api_call
from webapp import task_control

# In-memory progress registry for the streamed audit job. Safe under the app's
# single-process, multi-threaded gunicorn deployment (see task_control.py for
# the same reasoning). Keyed by task_id.
audit_progress_store = {}

# Extension types that represent an actual person (a "user"). These are the
# only extensions whose regional timezone is meaningful to compare against the
# site they belong to.
USER_TYPES = {'User', 'DigitalUser', 'FlexibleUser'}

# The exact column order used for both the audit export and the update upload.
# Keeping them identical means the file the operator downloads is the same file
# they edit and upload back.
TEMPLATE_COLUMNS = [
    "User Name",
    "Extension Number",
    "Site",
    "Site Timezone",
    "User Timezone",
]


# ---------------------------------------------------------------------------
# Directory fetching
# ---------------------------------------------------------------------------

def _paged(endpoint, token):
    """Yields every record across all pages of a collection endpoint."""
    page = 1
    while True:
        sep = "&" if "?" in endpoint else "?"
        resp = rc_api_call(
            f"{endpoint}{sep}perPage=1000&page={page}",
            token=token, raise_error=False
        )
        if not resp or 'records' not in resp:
            break
        for rec in resp['records']:
            yield rec
        if not resp.get('navigation', {}).get('nextPage'):
            break
        page += 1
        time.sleep(0.05)


def fetch_user_extensions(token):
    """Returns the list of user-type extensions (basic list records)."""
    return [e for e in _paged("/restapi/v1.0/account/~/extension", token)
            if e.get('type') in USER_TYPES and e.get('id')]


def build_site_timezone_map(token):
    """Returns {site_id (str): {'name': ..., 'tz_name': ...}}.

    The account's primary site is keyed under the well-known 'main-site' alias
    (and its real id, when known) using the account-level regional settings,
    so users on the Main Site can always be resolved.
    """
    site_map = {}
    for s in _paged("/restapi/v1.0/account/~/sites", token):
        sid = str(s.get('id'))
        tz = s.get('regionalSettings', {}).get('timezone')
        # The /sites collection does not always include regionalSettings, so
        # fetch the site detail when the timezone is missing.
        if not tz:
            detail = rc_api_call(
                f"/restapi/v1.0/account/~/sites/{sid}",
                token=token, raise_error=False
            )
            if detail:
                tz = detail.get('regionalSettings', {}).get('timezone')
            time.sleep(0.05)
        site_map[sid] = {
            'name': s.get('name', ''),
            'tz_name': str((tz or {}).get('name', '')),
        }

    # Account-level regional settings represent the primary ("Main") site,
    # which is not always returned by the /sites collection.
    account = rc_api_call("/restapi/v1.0/account/~", token=token, raise_error=False)
    if account:
        acc_tz = account.get('regionalSettings', {}).get('timezone')
        site_map.setdefault('main-site', {
            'name': 'Main Site',
            'tz_name': str((acc_tz or {}).get('name', '')),
        })

    return site_map


def _resolve_site(site_obj, site_map):
    """Maps an extension's site object to (site_name, site_tz_name)."""
    site_obj = site_obj or {}
    site_id = str(site_obj.get('id', '')) or 'main-site'
    info = site_map.get(site_id)
    if info is None and site_id != 'main-site':
        info = site_map.get('main-site')
    info = info or {}
    name = info.get('name') or site_obj.get('name', '') or 'Main Site'
    return name, info.get('tz_name', '')


# ---------------------------------------------------------------------------
# Audit (streamed background job)
# ---------------------------------------------------------------------------

def _build_audit_workbook(rows):
    """Renders audit rows to an .xlsx BytesIO, mismatches sorted to the top and
    highlighted in red."""
    df = pd.DataFrame(rows, columns=TEMPLATE_COLUMNS)

    if not df.empty:
        mismatch = df["Site Timezone"].astype(str).str.strip() != df["User Timezone"].astype(str).str.strip()
        df["_sort"] = (~mismatch).astype(int)  # mismatches (False->0) first
        df.sort_values(by=["_sort", "Site", "Extension Number"], inplace=True)
        df.drop(columns=["_sort"], inplace=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='User Timezone Audit')
        ws = writer.sheets['User Timezone Audit']

        for column in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(length + 5, 50)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        headers = [c.value for c in ws[1]]
        site_col = headers.index("Site Timezone") + 1
        user_col = headers.index("User Timezone") + 1
        for row in range(2, ws.max_row + 1):
            site_tz = str(ws.cell(row=row, column=site_col).value or "").strip()
            user_tz = str(ws.cell(row=row, column=user_col).value or "").strip()
            if site_tz != user_tz:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red

    output.seek(0)
    return output.getvalue()


def run_timezone_audit(task_id, token):
    """Background worker: builds the per-user timezone-vs-site audit and stores
    the resulting workbook in ``audit_progress_store`` for download."""
    audit_progress_store[task_id] = {
        'current': 0, 'total': 0, 'status': 'running', 'file_ready': False, 'mismatches': 0
    }
    try:
        site_map = build_site_timezone_map(token)
        users = fetch_user_extensions(token)

        store = audit_progress_store[task_id]
        store['total'] = len(users)

        rows = []
        mismatches = 0
        for idx, ext in enumerate(users):
            if task_control.is_stopped(task_id):
                store['status'] = 'cancelled'
                task_control.clear(task_id)
                return

            detail = rc_api_call(
                f"/restapi/v1.0/account/~/extension/{ext['id']}",
                token=token, raise_error=False
            )
            time.sleep(0.05)
            store['current'] = idx + 1

            if not detail:
                continue

            user_tz_name = str(detail.get('regionalSettings', {}).get('timezone', {}).get('name', ''))
            site_name, site_tz_name = _resolve_site(detail.get('site'), site_map)

            if user_tz_name.strip() != site_tz_name.strip():
                mismatches += 1

            rows.append({
                "User Name": detail.get('name', ''),
                "Extension Number": str(detail.get('extensionNumber', '')),
                "Site": site_name,
                "Site Timezone": site_tz_name,
                "User Timezone": user_tz_name,
            })

        store['mismatches'] = mismatches
        store['file_data'] = _build_audit_workbook(rows)
        store['status'] = 'completed'
        store['file_ready'] = True
    except Exception as e:
        audit_progress_store[task_id]['status'] = 'error'
        audit_progress_store[task_id]['error'] = str(e)


# ---------------------------------------------------------------------------
# Update (streamed upload -> preview / apply)
# ---------------------------------------------------------------------------

def _load_timezone_dictionary(token):
    """Returns (records, name_to_id) for the RC timezone dictionary."""
    records = list(_paged("/restapi/v1.0/dictionary/timezone", token))
    name_to_id = {}
    for tz in records:
        name_to_id[str(tz.get('name', '')).lower().strip()] = str(tz.get('id'))
        name_to_id[str(tz.get('id'))] = str(tz.get('id'))
    return records, name_to_id


def _resolve_timezone_id(raw, name_to_id, records):
    """Resolves a user-supplied timezone string to a RC timezone id, tolerant of
    the canonical name ('Australia/Sydney'), the id, or a city/description."""
    raw_lower = str(raw).lower().strip()
    if not raw_lower:
        return None
    if raw_lower in name_to_id:
        return name_to_id[raw_lower]
    city = raw_lower.split('/')[-1].replace('_', ' ')
    for tz in records:
        if city and (city in str(tz.get('name', '')).lower()
                     or city in str(tz.get('description', '')).lower()):
            return str(tz['id'])
    return None


def _get(row, key):
    val = row.get(key, '')
    if val is None:
        return ''
    return str(val).strip()


def update_timezone_batch(records, token, is_preview=False, task_id=None):
    """Generator yielding NDJSON-friendly progress dicts while previewing or
    applying user timezone changes from an uploaded audit file."""
    total = len(records)
    yield {"type": "start", "total": total,
           "message": "Loading account directories..."}

    # Verify the token and load supporting directories.
    account = rc_api_call("/restapi/v1.0/account/~", token=token, raise_error=False)
    if not account:
        yield {"type": "error", "message": "Unauthorized. Token expired or invalid."}
        return

    tz_records, name_to_id = _load_timezone_dictionary(token)
    if not tz_records:
        yield {"type": "error", "message": "Failed to load the timezone dictionary."}
        return

    # Build an extension-number -> id map for user extensions.
    ext_map = {}
    for e in _paged("/restapi/v1.0/account/~/extension", token):
        num = str(e.get('extensionNumber', '')).strip()
        if num and e.get('id'):
            ext_map[num] = str(e['id'])

    updated = skipped = errors = 0

    for i, row in enumerate(records):
        if not is_preview and task_control.is_stopped(task_id):
            yield {"type": "cancelled", "current": i, "total": total,
                   "message": f"Stopped by user. {i} of {total} row(s) processed; the rest were skipped."}
            task_control.clear(task_id)
            return

        name = _get(row, "User Name")
        ext_num = _get(row, "Extension Number").split('.')[0]
        desired_raw = _get(row, "User Timezone")

        def emit(status, message, from_tz="", to_tz=""):
            return {
                "type": "row", "current": i + 1, "total": total,
                "is_preview": is_preview, "name": name, "ext": ext_num,
                "from": from_tz, "to": to_tz, "status": status, "message": message,
            }

        if not ext_num:
            errors += 1
            yield emit("error", "Missing Extension Number.")
            continue

        ext_id = ext_map.get(ext_num)
        if not ext_id:
            errors += 1
            yield emit("error", f"No user extension found for number {ext_num}.")
            continue

        new_tz_id = _resolve_timezone_id(desired_raw, name_to_id, tz_records)
        if not new_tz_id:
            errors += 1
            yield emit("error", f"Unrecognised timezone: '{desired_raw}'.")
            continue

        detail = rc_api_call(
            f"/restapi/v1.0/account/~/extension/{ext_id}",
            token=token, raise_error=False
        )
        if not detail:
            errors += 1
            yield emit("error", "Failed to read the current extension settings.")
            continue

        cur_tz = detail.get('regionalSettings', {}).get('timezone', {}) or {}
        cur_tz_id = str(cur_tz.get('id', ''))
        cur_tz_name = str(cur_tz.get('name', ''))
        new_tz_name = next((str(t.get('name', '')) for t in tz_records
                            if str(t.get('id')) == new_tz_id), desired_raw)

        if cur_tz_id == new_tz_id:
            skipped += 1
            yield emit("skipped", "Already set — no change.", cur_tz_name, new_tz_name)
            continue

        if is_preview:
            yield emit("pending", "Will update.", cur_tz_name, new_tz_name)
            continue

        payload = {"regionalSettings": {"timezone": {"id": new_tz_id}}}
        resp = rc_api_call(
            f"/restapi/v1.0/account/~/extension/{ext_id}",
            method="PUT", json=payload, token=token, return_response=True
        )
        time.sleep(0.05)
        if resp is not None and getattr(resp, 'ok', False):
            updated += 1
            yield emit("applied", "Timezone updated.", cur_tz_name, new_tz_name)
        else:
            errors += 1
            err_text = ""
            try:
                err_text = str(resp.json().get('message', '')) if resp is not None else ""
            except Exception:
                err_text = getattr(resp, 'text', '') if resp is not None else ""
            yield emit("error", f"Update failed. {err_text}".strip(), cur_tz_name, new_tz_name)

    if not is_preview:
        task_control.clear(task_id)

    yield {"type": "done", "is_preview": is_preview,
           "updated": updated, "skipped": skipped, "errors": errors,
           "total": total}
