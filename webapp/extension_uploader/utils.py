import io
import time

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from webapp.rc_api import rc_api_call
from webapp import task_control


# ---------------------------------------------------------------------------
# User Type -> RingCentral extension `type` mapping
# ---------------------------------------------------------------------------
# The friendly User Type shown in the template maps to the RingCentral API
# extension `type` used when POSTing a new extension. These were confirmed with
# the operator; every row is still validated live against the account during
# preview, so any type the account rejects surfaces before anything is created.
USER_TYPE_MAP = {
    'User': 'User',
    'Virtual User': 'VirtualUser',
    'Message Only': 'Voicemail',
    'Announcement Only': 'Announcement',
    'Limited Extension': 'Limited',
}

# Ordered fallback API types tried when the account's service plan rejects the
# primary type with "Invalid user type [X] for current account service plan".
# RingCentral exposes several equivalent "user" types (User, DigitalUser,
# VirtualUser, FlexibleUser) and which ones a createExtension is allowed to use
# depends on the account's service plan. In particular the newer flexible plans
# expose the video-capable, no-digital-line user (shown as "Video Pro" in the
# portal) as FlexibleUser rather than the legacy VirtualUser. When a plan rejects
# the primary type, the create/validate is retried with each fallback in turn and
# the substitution is surfaced to the operator so they can confirm the created
# extension in the portal.
USER_TYPE_FALLBACKS = {
    'VirtualUser': ['FlexibleUser'],
}

# The template's dropdown order for the User Type column.
USER_TYPES = ['User', 'Virtual User', 'Message Only', 'Announcement Only', 'Limited Extension']

# For these types only the First Name is used -- Last Name is intentionally
# ignored (RingCentral stores these as single-name extensions).
FIRST_NAME_ONLY_TYPES = {'Message Only', 'Announcement Only', 'Limited Extension'}

# The spreadsheet column order (also the sheet header row).
# 'Email' is the extension's contact email (set at create time). 'Notification
# Email' is the address that receives voicemail/missed-call notifications and can
# differ from the contact email; it is applied in a second pass after create.
TEMPLATE_HEADERS = [
    'Site Name', 'Extension', 'First Name', 'Last Name',
    'Email', 'User Type', 'Department', 'Role', 'Notification Email',
]

TEMPLATE_SHEET = 'Extension Upload'
# How many rows the dropdown data-validations cover in the blank template.
TEMPLATE_VALIDATION_ROWS = 500


# ---------------------------------------------------------------------------
# Directory / reference fetching
# ---------------------------------------------------------------------------

def fetch_all_extensions(token):
    """Fetches every account extension (all pages). Used to detect extension
    numbers and emails that are already in use."""
    extensions = []
    page = 1
    while True:
        resp = rc_api_call(
            f"/restapi/v1.0/account/~/extension?perPage=1000&page={page}",
            token=token, raise_error=False
        )
        if not resp or 'records' not in resp:
            break
        extensions.extend(resp['records'])
        if not resp.get('navigation', {}).get('nextPage'):
            break
        page += 1
        time.sleep(0.05)
    return extensions


def fetch_sites(token):
    """Fetches every Site on the account (all pages).

    Returns a list of {'id', 'name'}. The account's primary site is always
    represented as the well-known 'main-site' id / 'Main Site' name so it can be
    selected even though it is not always returned by the /sites collection."""
    sites = []
    page = 1
    while True:
        resp = rc_api_call(
            f"/restapi/v1.0/account/~/sites?perPage=1000&page={page}",
            token=token, raise_error=False
        )
        if not resp or 'records' not in resp:
            break
        for s in resp['records']:
            sites.append({'id': s.get('id'), 'name': s.get('name', '')})
        if not resp.get('navigation', {}).get('nextPage'):
            break
        page += 1
        time.sleep(0.05)

    if not any((s.get('id') == 'main-site') for s in sites):
        sites.insert(0, {'id': 'main-site', 'name': 'Main Site'})
    return sites


def fetch_roles(token):
    """Fetches every assignable user role on the account (all pages).

    Returns a list of {'id', 'name'} using the role's friendly displayName."""
    roles = []
    page = 1
    while True:
        resp = rc_api_call(
            f"/restapi/v1.0/account/~/user-role?perPage=1000&page={page}",
            token=token, raise_error=False
        )
        if not resp or 'records' not in resp:
            break
        for r in resp['records']:
            name = r.get('displayName') or r.get('name') or ''
            roles.append({'id': r.get('id'), 'name': name})
        if not resp.get('navigation', {}).get('nextPage'):
            break
        page += 1
        time.sleep(0.05)
    return roles


# ---------------------------------------------------------------------------
# Lookups
# ---------------------------------------------------------------------------

def build_site_lookup(sites):
    """Maps a friendly Site name (case-insensitive) -> Site id."""
    lookup = {}
    for s in sites:
        name = (s.get('name') or '').strip().lower()
        if name:
            lookup[name] = s.get('id')
    lookup.setdefault('main site', 'main-site')
    return lookup


def build_role_lookup(roles):
    """Maps a friendly role name (case-insensitive) -> role id."""
    lookup = {}
    for r in roles:
        name = (r.get('name') or '').strip().lower()
        if name:
            lookup[name] = r.get('id')
    return lookup


def build_existing_lookups(extensions):
    """Returns (numbers, emails) already in use on the account.

    numbers: {extensionNumber(str) -> display name}
    emails:  {email(lower) -> display name}
    """
    numbers = {}
    emails = {}
    for ext in extensions:
        num = ext.get('extensionNumber')
        display = ext.get('name') or f"Ext {num}"
        if num is not None:
            numbers[str(num).strip()] = display
        contact = ext.get('contact') or {}
        email = (contact.get('email') or '').strip().lower()
        if email:
            emails[email] = display
    return numbers, emails


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------

def generate_template(token):
    """Builds the blank Extension Uploader workbook.

    Sheet 1 ('Extension Upload') carries the eight upload columns with dropdown
    validation on Site Name, User Type and Role. Sheet 2 ('Reference') holds the
    live Site and Role lists (polled from the connected account) that the
    dropdowns point at, plus the fixed User Type list."""
    sites = fetch_sites(token)
    roles = fetch_roles(token)
    site_names = sorted({(s.get('name') or '').strip() for s in sites if s.get('name')})
    role_names = sorted({(r.get('name') or '').strip() for r in roles if r.get('name')})

    wb = Workbook()
    ws = wb.active
    ws.title = TEMPLATE_SHEET
    ws.append(TEMPLATE_HEADERS)

    # Reference sheet the dropdowns point at (range refs avoid the 255-char
    # inline-list limit that long Site / Role lists would blow past).
    ref = wb.create_sheet('Reference')
    ref['A1'] = 'User Types'
    ref['B1'] = 'Sites'
    ref['C1'] = 'Roles'
    for i, v in enumerate(USER_TYPES, start=2):
        ref.cell(row=i, column=1, value=v)
    for i, v in enumerate(site_names, start=2):
        ref.cell(row=i, column=2, value=v)
    for i, v in enumerate(role_names, start=2):
        ref.cell(row=i, column=3, value=v)

    last = TEMPLATE_VALIDATION_ROWS + 1

    # User Type dropdown (column F).
    ut_dv = DataValidation(
        type='list',
        formula1=f"=Reference!$A$2:$A${1 + len(USER_TYPES)}",
        allow_blank=False,
    )
    ws.add_data_validation(ut_dv)
    ut_dv.add(f"F2:F{last}")

    # Site Name dropdown (column A).
    if site_names:
        site_dv = DataValidation(
            type='list',
            formula1=f"=Reference!$B$2:$B${1 + len(site_names)}",
            allow_blank=True,
        )
        ws.add_data_validation(site_dv)
        site_dv.add(f"A2:A{last}")

    # Role dropdown (column H).
    if role_names:
        role_dv = DataValidation(
            type='list',
            formula1=f"=Reference!$C$2:$C${1 + len(role_names)}",
            allow_blank=True,
        )
        ws.add_data_validation(role_dv)
        role_dv.add(f"H2:H{last}")

    # Column widths for readability.
    widths = {'A': 22, 'B': 12, 'C': 16, 'D': 16, 'E': 30, 'F': 20, 'G': 18, 'H': 22, 'I': 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ref.column_dimensions['A'].width = 20
    ref.column_dimensions['B'].width = 24
    ref.column_dimensions['C'].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Cell cleaning
# ---------------------------------------------------------------------------

def _clean(value):
    """Normalises a spreadsheet cell to a trimmed string (handles NaN / '1001.0')."""
    text = str(value).strip()
    if text.lower() == 'nan':
        return ''
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    return text


# ---------------------------------------------------------------------------
# Extension creation
# ---------------------------------------------------------------------------

def _error_message(resp):
    """Human-readable error string from an RC response, preserving RC's message."""
    if resp is None:
        return 'No response from RingCentral'
    try:
        body = resp.json() or {}
    except Exception:
        body = {}
    msg = body.get('message') or body.get('error')
    if not msg and isinstance(body.get('errors'), list) and body['errors']:
        msg = body['errors'][0].get('message')
    if not msg:
        msg = getattr(resp, 'text', '') or f"HTTP {getattr(resp, 'status_code', '?')}"
    return str(msg)[:300]


def _full_error(resp):
    """Verbose error string for create/update failures.

    Unlike `_error_message` (which returns only the top-level message), this
    surfaces the HTTP status, RC's `errorCode`, and any nested per-field
    `errors[]` detail -- so a bare 500 "Internal Server Error" still reveals the
    specific RC code/reason hidden underneath it, which is exactly what we need
    to see for the Message Only (Voicemail) create failures."""
    if resp is None:
        return 'No response from RingCentral'
    status = getattr(resp, 'status_code', '?')
    try:
        body = resp.json() or {}
    except Exception:
        body = {}
    code = body.get('errorCode')
    msg = body.get('message') or body.get('error') or ''
    nested = []
    errs = body.get('errors')
    if isinstance(errs, list):
        for e in errs[:3]:
            if not isinstance(e, dict):
                continue
            ec = e.get('errorCode')
            em = e.get('message') or ''
            piece = f"{ec + ': ' if ec else ''}{em}".strip()
            if piece:
                nested.append(piece)
    head = f"HTTP {status}"
    if code:
        head += f" · {code}"
    tail = '; '.join(nested) or msg or (getattr(resp, 'text', '') or '')
    out = f"{head} — {tail}".strip(' —').strip()
    return out[:400]


def _rc_write(endpoint, payload, token, method='POST'):
    """POST/PUT helper with a short 429 back-off loop.

    Returns (ok, body_or_none, response)."""
    for _ in range(4):
        resp = rc_api_call(endpoint, method=method, json=payload, token=token, return_response=True)
        status_code = getattr(resp, 'status_code', None)
        if status_code == 429:
            retry_after = int(resp.headers.get('Retry-After', 10)) if hasattr(resp, 'headers') else 10
            time.sleep(retry_after + 1)
            continue
        if resp is not None and getattr(resp, 'ok', False):
            try:
                return True, (resp.json() or {}), resp
            except Exception:
                return True, {}, resp
        return False, None, resp
    return False, None, None


def _apply_notification_settings(ext_id, email, enable_transcription, token):
    """Read-modify-write the extension's notification-settings.

    Sets the voicemail/notification recipient email (when given), and for
    Message Only (Voicemail) extensions enables Voicemail-to-Text by turning on
    email notification + transcription on the voicemail channel -- mirroring the
    account's cq_hours notifications updater.

    A bare {"emailAddresses": [...]} PUT makes RC validate every channel, and a
    Message Only box has no inbound SMS / fax / missed-call channels, so it
    answers 400 CMN-101 "Parameter [inboundTexts]/[inboundFaxes]/[missedCalls]
    value is invalid". Reading the current settings first (they come back shaped
    for this extension's type), dropping the read-only `emailRecipients` echo,
    and writing the same structure back keeps only the valid channels. Accounts
    without the Voicemail-to-Text feature reject the transcription toggles, so
    those are dropped and retried. Returns (ok, response, transcription_dropped).
    """
    endpoint = f'/restapi/v1.0/account/~/extension/{ext_id}/notification-settings'
    current = rc_api_call(endpoint, token=token, raise_error=False)
    if isinstance(current, dict):
        settings = dict(current)
        settings.pop('emailRecipients', None)
    else:
        # Couldn't read current settings -- best-effort minimal body.
        settings = {}

    if email:
        # In advanced mode each channel has its own recipients; otherwise the
        # top-level list is shared (same split as the cq_hours updater).
        if settings.get('advancedMode'):
            vm = dict(settings.get('voicemails') or {})
            vm['emailAddresses'] = [email]
            settings['voicemails'] = vm
        else:
            settings['emailAddresses'] = [email]

    if enable_transcription:
        vm = dict(settings.get('voicemails') or {})
        vm['notifyByEmail'] = True
        vm['includeTranscription'] = True
        settings['voicemails'] = vm

    ok, _b, resp = _rc_write(endpoint, settings, token, method='PUT')

    # Accounts without Voicemail-to-Text reject the transcription (and related)
    # toggles -- drop them and retry so the email still applies (cq_hours does
    # the same).
    transcription_dropped = False
    if not ok and enable_transcription and isinstance(settings.get('voicemails'), dict):
        _code, msg = _error_code_and_message(resp)
        if any(k in (msg or '') for k in ('includeTranscription', 'includeAttachment', 'markAsRead')):
            vm = dict(settings['voicemails'])
            vm.pop('includeTranscription', None)
            vm.pop('includeAttachment', None)
            vm.pop('markAsRead', None)
            settings['voicemails'] = vm
            r_ok, _rb, r_resp = _rc_write(endpoint, settings, token, method='PUT')
            if r_ok:
                ok, resp, transcription_dropped = True, r_resp, True
            else:
                resp = r_resp
    return ok, resp, transcription_dropped


def _create_one(api_type, contact, ext_number, site_id, token):
    """POST a single extension of one specific api_type.

    Handles the main-site quirk and the drop-site retry (both independent of the
    user type). Returns (ok, body, resp, site_dropped)."""
    base_payload = {
        'type': api_type,
        'extensionNumber': ext_number,
        'contact': contact,
    }
    # The synthetic 'main-site' token is exactly what createExtension rejects
    # with an opaque 500, and new extensions default to the main site when no
    # site is sent -- so for main-site rows omit the site up front (this avoids
    # a guaranteed-failing first POST). Real sites use a numeric id and are sent
    # normally; the retry-without-site below is their safety net.
    send_site = bool(site_id) and str(site_id) != 'main-site'
    payload = dict(base_payload)
    if send_site:
        payload['site'] = {'id': str(site_id)}

    ok, body, resp = _rc_write('/restapi/v1.0/account/~/extension', payload, token, method='POST')

    # A real (numeric-id) site can still be rejected with a bare 500 -- e.g. a
    # site the type can't live on. The number itself is fine (it's in the type's
    # free-number pool and validate passed), so retry once without the site; the
    # extension then lands on the account's main site, flagged with a warning.
    # Don't retry the genuine "no numbers" (EXT-250) or "wrong type for plan"
    # conditions -- neither is a site problem, and dropping the site won't help.
    site_dropped = False
    if not ok and send_site:
        code, msg = _error_code_and_message(resp)
        if not _looks_like_no_number(code, msg) and not _looks_like_invalid_type_for_plan(code, msg):
            r_ok, r_body, r_resp = _rc_write(
                '/restapi/v1.0/account/~/extension', base_payload, token, method='POST')
            if r_ok:
                ok, body, resp, site_dropped = True, r_body, r_resp, True

    return ok, body, resp, site_dropped


def create_extension(plan, token):
    """Creates one extension from a validated plan dict, then assigns its role if
    one was resolved. Returns (ok, message)."""
    contact = {'firstName': plan['first_name']}
    if plan.get('last_name') and not plan['drop_last_name']:
        contact['lastName'] = plan['last_name']
    if plan.get('email'):
        contact['email'] = plan['email']
    if plan.get('department'):
        contact['department'] = plan['department']

    # `contact.firstName` and `contact.email` are the only contact fields the
    # createExtension schema (ContactInfoCreationRequest) marks required.
    #
    # Try the plan's primary type, then any service-plan fallbacks: if RC rejects
    # the type as unsupported by the account's plan ("Invalid user type [X] for
    # current account service plan") we retry with an equivalent type the plan
    # does support. Any other failure stops the chain and is reported as-is.
    api_type = plan['api_type']
    site_id = plan.get('site_id')
    candidate_types = [api_type] + USER_TYPE_FALLBACKS.get(api_type, [])
    ok = False
    body = resp = None
    site_dropped = False
    used_type = api_type
    for candidate in candidate_types:
        ok, body, resp, site_dropped = _create_one(candidate, contact, plan['ext'], site_id, token)
        if ok:
            used_type = candidate
            break
        code, msg = _error_code_and_message(resp)
        if not _looks_like_invalid_type_for_plan(code, msg):
            break

    if not ok:
        return False, f"Create failed — {_full_error(resp)}"

    new_id = (body or {}).get('id')

    # The role and the notification email cannot be set on the create payload, so
    # they are applied in follow-up passes. A failure in either leaves the
    # extension created, so it is reported as a partial success (a warning), not
    # a hard failure.
    warnings = []
    if used_type != api_type:
        warnings.append(
            f"created as {used_type} — the account's service plan does not support "
            f"{api_type} (please confirm the extension in the portal)")
    if site_dropped:
        warnings.append("created on the main site (RingCentral rejected the site assignment)")

    if plan.get('role_id') and new_id:
        r_ok, _r_body, r_resp = _rc_write(
            f'/restapi/v1.0/account/~/extension/{new_id}/assigned-role',
            {'records': [{'id': str(plan['role_id'])}]},
            token, method='PUT',
        )
        if not r_ok:
            warnings.append(f"role assignment failed ({_full_error(r_resp)})")

    # Second pass: set the notification email (voicemail / missed-call notices).
    # This is independent of the contact email set at create time and may differ.
    # Message Only (Voicemail) extensions are created with Voicemail-to-Text
    # enabled. The notification email (if given) is applied in the same pass.
    enable_vm2t = plan.get('api_type') == 'Voicemail'
    if (plan.get('notif_email') or enable_vm2t) and new_id:
        n_ok, n_resp, vm2t_dropped = _apply_notification_settings(
            new_id, plan.get('notif_email'), enable_vm2t, token)
        if not n_ok:
            labels = []
            if plan.get('notif_email'):
                labels.append('notification email')
            if enable_vm2t:
                labels.append('voicemail-to-text')
            warnings.append(f"{' / '.join(labels)} update failed ({_full_error(n_resp)})")
        elif vm2t_dropped:
            warnings.append("voicemail-to-text not enabled (account feature unavailable)")

    if warnings:
        return True, 'Created, but ' + '; '.join(warnings)
    return True, 'Created'


# ---------------------------------------------------------------------------
# Preflight (review-time checks against the live account)
# ---------------------------------------------------------------------------
#
# EXT-250 "There are no extension numbers available" is raised by createExtension
# when the requested extension number isn't an assignable free number for that
# type (empty pool, out-of-range, or reserved). These two Provisioning-API calls
# surface that at review time instead of at Create:
#   GET  extension/free-numbers?extensionType=<type>  -> is there ANY free number?
#   POST extension/validate                            -> is THIS number creatable?
# Both are public ReadAccounts/EditAccounts endpoints, so the tool's own token
# can call them.

# Sentinel so a cached "we couldn't determine capacity" (None) is distinct from
# "not looked up yet".
_UNSET = object()


def _error_code_and_message(resp):
    """(errorCode, message) from an RC error body, preserving RC's own values."""
    if resp is None:
        return None, 'No response from RingCentral'
    try:
        body = resp.json() or {}
    except Exception:
        body = {}
    code = body.get('errorCode')
    msg = body.get('message') or body.get('error')
    errs = body.get('errors')
    if isinstance(errs, list) and errs:
        code = code or errs[0].get('errorCode')
        msg = msg or errs[0].get('message')
    if not msg:
        msg = getattr(resp, 'text', '') or f"HTTP {getattr(resp, 'status_code', '?')}"
    return code, str(msg)[:300]


def _looks_like_no_number(code, msg):
    """True if an RC error is the 'no available extension number' condition.

    We match narrowly (by code/message) so unrelated validation quirks -- e.g.
    the validate endpoint not recognising the 'Limited' type -- don't get
    misreported as a capacity problem."""
    c = (code or '').upper()
    m = (msg or '').lower()
    return (
        c == 'EXT-250'
        or 'no extension numbers available' in m
        or ('extension number' in m and 'available' in m)
    )


def _looks_like_invalid_type_for_plan(code, msg):
    """True if RC rejected the extension `type` as unsupported by the account's
    current service plan -- e.g. HTTP 400 "Invalid user type [VirtualUser] for
    current account service plan".

    This is a plan/type mismatch, not a bad-data or capacity problem, so the
    caller can safely retry with an equivalent user type the plan does support
    (see USER_TYPE_FALLBACKS). Matched on the message text so it works regardless
    of the (inconsistently spelled) errorCode RC returns alongside it."""
    m = (msg or '').lower()
    return (
        'invalid user type' in m
        or ('user type' in m and 'service plan' in m)
    )


def fetch_free_numbers(token, api_type, count=1):
    """Available free extension numbers for a given extension type.

    Returns a list (possibly empty) on success, or None if the endpoint could
    not be queried (e.g. the type isn't supported by this endpoint, like
    'Limited'), so the caller can degrade to 'unknown' rather than block."""
    endpoint = (
        f"/restapi/v1.0/account/~/extension/free-numbers"
        f"?extensionType={api_type}&count={count}"
    )
    for _ in range(3):
        resp = rc_api_call(endpoint, token=token, return_response=True)
        code = getattr(resp, 'status_code', None)
        if code == 429:
            wait = int(resp.headers.get('Retry-After', 5)) + 1 if hasattr(resp, 'headers') else 5
            time.sleep(wait)
            continue
        if resp is not None and getattr(resp, 'ok', False):
            try:
                body = resp.json() or {}
            except Exception:
                return None
            nums = body.get('extensionNumbers')
            return nums if isinstance(nums, list) else []
        return None
    return None


def _validate_one(api_type, plan, token):
    """POST extension/validate for one specific api_type. Returns the response
    (after a short 429 back-off), or None if it couldn't be called."""
    contact = {'firstName': plan['first_name']}
    if plan.get('last_name') and not plan['drop_last_name']:
        contact['lastName'] = plan['last_name']
    if plan.get('email'):
        contact['email'] = plan['email']
    if plan.get('department'):
        contact['department'] = plan['department']
    payload = {'type': api_type, 'extensionNumber': plan['ext'], 'contact': contact}
    if plan.get('site_id'):
        payload['site'] = {'id': str(plan['site_id'])}

    for _ in range(3):
        resp = rc_api_call('/restapi/v1.0/account/~/extension/validate', method='POST',
                           json=payload, token=token, return_response=True)
        code = getattr(resp, 'status_code', None)
        if code == 429:
            wait = int(resp.headers.get('Retry-After', 5)) + 1 if hasattr(resp, 'headers') else 5
            time.sleep(wait)
            continue
        return resp
    return None


def validate_new_extension(plan, token, invalid_types=None):
    """Dry-run an extension via POST extension/validate (creates nothing).

    Tries the plan's primary type then any service-plan fallbacks, so a type the
    account's plan rejects (e.g. VirtualUser -> FlexibleUser) is resolved here,
    at review time, exactly as Create will resolve it. `invalid_types` is an
    optional set reused across rows to remember which primary types this account's
    plan has already rejected, so each is validated at most once per batch.

    Returns (state, message, effective_type):
      ('ok', None, <type>)         -- RingCentral accepts this extension as <type>
      ('invalid', message, <type>) -- rejected for a *number availability* reason
                                      (EXT-250 etc.) -- the thing we want to flag
      ('badtype', message, None)   -- the primary type and every fallback were
                                      rejected as unsupported by the service plan
      ('unavailable', None, None)  -- couldn't validate, or the rejection was for
                                      an unrelated reason; caller should not block
    """
    api_type = plan['api_type']
    badtype_msg = None
    for candidate in [api_type] + USER_TYPE_FALLBACKS.get(api_type, []):
        if invalid_types is not None and candidate in invalid_types:
            badtype_msg = badtype_msg or f"Invalid user type [{candidate}] for current account service plan"
            continue
        resp = _validate_one(candidate, plan, token)
        if resp is not None and getattr(resp, 'ok', False):
            return 'ok', None, candidate
        err_code, msg = _error_code_and_message(resp)
        if _looks_like_invalid_type_for_plan(err_code, msg):
            if invalid_types is not None:
                invalid_types.add(candidate)
            badtype_msg = badtype_msg or msg
            continue  # this type isn't allowed on the plan -- try the next one
        if _looks_like_no_number(err_code, msg):
            return 'invalid', f"{err_code + ': ' if err_code else ''}{msg}", candidate
        return 'unavailable', None, None
    # Every candidate type was rejected as unsupported by the account's plan.
    if badtype_msg is not None:
        return 'badtype', badtype_msg, None
    return 'unavailable', None, None


def preflight_row(plan, token, free_cache, invalid_types=None):
    """Review-time check for one planned extension against the live account.

    Returns (ok, message). ok=False means the row would be rejected on Create
    (surfaced now instead); message is the reason. ok=True means it will be
    created; message is an optional note (e.g. a service-plan type substitution)
    or None. free_cache maps api_type -> free-number result and invalid_types
    remembers plan-rejected types, both reused across rows so each type is polled
    at most once per batch."""
    api_type = plan['api_type']

    # Confirm THIS specific number/type is creatable, resolving any service-plan
    # type substitution the same way Create will.
    state, vmsg, eff_type = validate_new_extension(plan, token, invalid_types)
    if state == 'invalid':
        return False, (
            f"RingCentral won't create extension {plan['ext']} — {vmsg}. "
            f"Try a different extension number within the account's range."
        )
    if state == 'badtype':
        tried = ', '.join([api_type] + USER_TYPE_FALLBACKS.get(api_type, []))
        return False, (
            f"RingCentral rejected the user type for this account's service plan "
            f"({vmsg}). Tried: {tried}. This plan may not support this user type."
        )
    if state == 'ok':
        note = None
        if eff_type and eff_type != api_type:
            note = (f"the account's service plan does not support {api_type}; "
                    f"it will be created as {eff_type} — please confirm in the portal")
        return True, note

    # validate couldn't give a definitive answer -- fall back to the free-number
    # heuristic so a definitively empty pool still surfaces EXT-250 at review time.
    cap = free_cache.get(api_type, _UNSET)
    if cap is _UNSET:
        cap = fetch_free_numbers(token, api_type, count=1)
        free_cache[api_type] = cap
    if cap == []:
        return False, (
            f"No available extension numbers for this type on the account "
            f"(RingCentral EXT-250). Check the account's extension-number range/capacity."
        )
    return True, None


# ---------------------------------------------------------------------------
# Validate / apply batch (streamed NDJSON)
# ---------------------------------------------------------------------------

def process_upload_batch(records, token, is_preview=True, task_id=None):
    """Validates (preview) or creates (apply) new extensions from the uploaded
    rows, yielding NDJSON-friendly progress chunks:

      {"type": "start", ...}
      {"type": "progress", "result": {...}, "is_preview": bool}
      {"type": "cancelled", ...}      (apply only, on Stop)
      {"type": "done", "is_preview": bool}

    Preview validates every row against the live account (extension free, email
    free, Site / Role / User Type resolvable) but writes nothing. Apply re-runs
    the same validation and POSTs each valid new extension."""
    total = len(records)
    yield {"type": "start", "total": total,
           "message": "Loading account directory (extensions, sites, roles)…"}

    try:
        extensions = fetch_all_extensions(token)
    except Exception as e:
        yield {"type": "error", "message": f"Failed to load extensions directory: {e}"}
        return
    if extensions is None:
        yield {"type": "error", "message": "Could not load account extensions. Token may be invalid or expired."}
        return

    try:
        sites = fetch_sites(token)
        roles = fetch_roles(token)
    except Exception as e:
        yield {"type": "error", "message": f"Failed to load sites / roles: {e}"}
        return

    site_lookup = build_site_lookup(sites)
    role_lookup = build_role_lookup(roles)
    existing_numbers, existing_emails = build_existing_lookups(extensions)

    # Detect collisions *within* the uploaded sheet too, so two rows that both
    # claim the same extension or email are flagged rather than silently racing.
    seen_ext = {}
    seen_email = {}

    # Free-number availability per extension type, polled once per type and
    # reused across rows during the review-time preflight.
    free_cache = {}
    # User types this account's service plan has rejected, remembered so each is
    # validated at most once per batch (see validate_new_extension).
    invalid_types = set()

    for i, row in enumerate(records):
        # Cooperative stop (apply only -- preview writes nothing).
        if not is_preview and task_control.is_stopped(task_id):
            yield {"type": "cancelled", "current": i, "total": total,
                   "message": f"Stopped by user. {i} of {total} row(s) processed; the rest were skipped."}
            task_control.clear(task_id)
            return

        site_name = _clean(row.get('Site Name', ''))
        ext = _clean(row.get('Extension', ''))
        first_name = _clean(row.get('First Name', ''))
        last_name = _clean(row.get('Last Name', ''))
        email = _clean(row.get('Email', ''))
        user_type = _clean(row.get('User Type', ''))
        department = _clean(row.get('Department', ''))
        role_name = _clean(row.get('Role', ''))
        notif_email = _clean(row.get('Notification Email', ''))

        # These types are single-name extensions: the Last Name column is
        # ignored. Email is still used for every type (e.g. voicemail notices).
        drop_last = user_type in FIRST_NAME_ONLY_TYPES

        def progress(status, message):
            return {
                "type": "progress",
                "current": i + 1,
                "total": total,
                "result": {
                    "row": i + 2,  # 1-based sheet row (header is row 1)
                    "ext": ext or "N/A",
                    "name": (first_name if drop_last else f"{first_name} {last_name}").strip() or "—",
                    "email": email,
                    "notif_email": notif_email,
                    "type": user_type or "—",
                    "site": site_name,
                    "role": role_name,
                    "status": status,
                    "message": message,
                },
                "is_preview": is_preview,
            }

        # Fully blank row -> skip silently.
        if not any([site_name, ext, first_name, last_name, email, user_type, department, role_name, notif_email]):
            yield progress("info", "Skipped empty row")
            continue

        # --- Required fields ---
        if not ext:
            yield progress("error", "Missing Extension")
            continue
        if not user_type:
            yield progress("error", "Missing User Type")
            continue
        if user_type not in USER_TYPE_MAP:
            yield progress("error", f"Unknown User Type '{user_type}'")
            continue
        if not first_name:
            yield progress("error", "Missing First Name")
            continue
        # Full users need an email; single-name types (message/announce/limited)
        # may have one but don't require it.
        if not drop_last and not email:
            yield progress("error", "Missing Email")
            continue

        # --- Duplicate extension (account or within-file) ---
        if ext in existing_numbers:
            yield progress("duplicate", f"Extension {ext} already in use on this account ({existing_numbers[ext]})")
            continue
        if ext in seen_ext:
            yield progress("duplicate", f"Extension {ext} appears more than once in this file (row {seen_ext[ext]})")
            continue

        # --- Duplicate email (account or within-file), whenever one is given ---
        email_l = email.lower()
        if email_l:
            if email_l in existing_emails:
                yield progress("duplicate", f"Email {email} already in use on this account ({existing_emails[email_l]})")
                continue
            if email_l in seen_email:
                yield progress("duplicate", f"Email {email} appears more than once in this file (row {seen_email[email_l]})")
                continue

        # --- Resolve Site (optional) ---
        site_id = None
        if site_name:
            site_id = site_lookup.get(site_name.lower())
            if not site_id:
                yield progress("error", f"Site '{site_name}' not found on this account")
                continue

        # --- Resolve Role (optional) ---
        role_id = None
        if role_name:
            role_id = role_lookup.get(role_name.lower())
            if not role_id:
                yield progress("error", f"Role '{role_name}' not found on this account")
                continue

        # Row is valid -- reserve its extension / email so later rows collide.
        seen_ext[ext] = i + 2
        if email_l:
            seen_email[email_l] = i + 2

        plan = {
            'ext': ext,
            'first_name': first_name,
            'last_name': last_name,
            'drop_last_name': drop_last,
            'email': email,
            'department': department,
            'user_type': user_type,
            'api_type': USER_TYPE_MAP[user_type],
            'site_id': site_id,
            'role_id': role_id,
            'notif_email': notif_email,
        }

        note = []
        if site_id:
            note.append(f"site → {site_name}")
        if role_id:
            note.append(f"role → {role_name}")
        if notif_email:
            note.append(f"notification → {notif_email}")
        if drop_last and last_name:
            note.append("last name ignored for this type")
        detail = f" ({'; '.join(note)})" if note else ""

        if is_preview:
            # Preflight against the live account so a number RingCentral would
            # reject (EXT-250) or a user type its service plan doesn't support
            # (VirtualUser -> FlexibleUser) is flagged here, at review time,
            # rather than only when Create is clicked.
            ok, pf_msg = preflight_row(plan, token, free_cache, invalid_types)
            if not ok:
                yield progress("error", pf_msg)
                continue
            extra = f" — {pf_msg}" if pf_msg else ""
            yield progress("success", f"Will create {user_type} ext {ext}{detail}{extra}")
            time.sleep(0.05)
            continue

        # Apply mode -- create the extension.
        ok, msg = create_extension(plan, token)
        yield progress("success" if ok else "error", msg)
        time.sleep(0.1)

    task_control.clear(task_id)
    yield {"type": "done", "is_preview": is_preview}
