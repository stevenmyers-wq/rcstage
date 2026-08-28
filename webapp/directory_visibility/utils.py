"""Company Directory Visibility — business logic and RingCentral calls.

RingCentral extensions carry a boolean ``hidden`` flag that controls whether the
object appears in the company (dial-by-name) directory. ``hidden: false`` means
the object is *Visible* in the directory; ``hidden: true`` means it is *Hidden*.

Not every extension type supports the flag — RingCentral only returns / accepts
``hidden`` for the object types that can live in the directory (the User-family
types). This module polls the account, surfaces the current Visible / Hidden
state for every object that supports it, and (because the updateExtension API
accepts the flag) offers a single-object toggle and a bulk XLSX-driven update.
"""
import io
import time

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from webapp.rc_api import rc_api_call
from webapp import task_control


# ---------------------------------------------------------------------------
# Which objects support company-directory visibility
# ---------------------------------------------------------------------------
# Two different scopes, deliberately kept apart:
#
#  * AUDIT (GET) — RingCentral exposes the `hidden` flag on the ExtensionInfo of
#    every object type that can appear in the company directory (Users, Call
#    Queues / Departments, IVR menus, …). So the read-only audit surfaces any
#    extension whose GET payload carries `hidden`, whatever its type.
#
#  * MODIFY (PUT) — updateExtension only accepts a change to `hidden` for User-
#    type extensions. Everything else is read-only via the API and must be
#    changed in the Admin Portal. The USER_LIKE_TYPES set below is the editable
#    subset (the User family exposed across service plans); rows outside it are
#    audited but flagged non-editable, and no PUT is attempted for them.
#
# (The portal-effective directory can also be read via
# GET /restapi/v1.0/account/~/directory/contacts, which simply omits anything
# hidden. This tool audits the flag itself so it can report the hidden objects
# too, not just the visible ones.)
USER_LIKE_TYPES = {'User', 'DigitalUser', 'VirtualUser', 'FlexibleUser'}

# Some extension types carry a `hidden` field in their payload but are not
# company-directory contacts, so the flag is meaningless for them and reads as a
# spurious (usually "Hidden") state — Site being the reported case (a Site is a
# location container, never a directory entry). Exclude these from the audit
# entirely rather than surface a misleading state.
NON_DIRECTORY_TYPES = {'Site'}

# Bulk template.
TEMPLATE_SHEET = 'Directory Visibility'
VISIBLE = 'Visible'
HIDDEN = 'Hidden'
VISIBILITY_CHOICES = [VISIBLE, HIDDEN]
TEMPLATE_HEADERS = [
    'Extension ID', 'Extension', 'Name', 'Type', 'Editable (API)',
    'Current Visibility', 'New Visibility',
]


# ---------------------------------------------------------------------------
# Directory fetching
# ---------------------------------------------------------------------------

def fetch_all_extensions(token):
    """Fetch every account extension (all pages) as raw records."""
    extensions = []
    page = 1
    while True:
        resp = rc_api_call(
            f"/restapi/v1.0/account/~/extension?perPage=1000&page={page}",
            token=token, raise_error=False
        )
        if not resp or 'records' not in resp:
            break
        extensions.extend(resp['records'])
        if not resp.get('navigation', {}).get('nextPage'):
            break
        page += 1
        time.sleep(0.05)
    return extensions


def _display_name(record):
    """Best display name for an extension record (contact name, then name)."""
    contact = record.get('contact') or {}
    first = (contact.get('firstName') or '').strip()
    last = (contact.get('lastName') or '').strip()
    combined = f"{first} {last}".strip()
    return combined or (record.get('name') or '').strip() or '—'


def build_visibility_rows(token):
    """Poll the account and return every object that exposes company-directory
    visibility, with its current Visible / Hidden state.

    Returns (rows, summary) where each row is::

        {id, extensionNumber, name, type, status,
         hidden(bool), visible(bool), editable(bool)}

    An object is included in the audit when RingCentral returns the `hidden`
    field for it (GET exposes it for any directory-capable type). ``editable`` is
    True only for User-type extensions — the ones whose visibility can be changed
    via the API; other types are read-only here and must be changed in the Admin
    Portal. summary counts visible / hidden / total plus editable / portal_only.
    """
    extensions = fetch_all_extensions(token)
    if extensions is None:
        return None, None

    rows = []
    for ext in extensions:
        # Audit any object whose GET payload carries the flag, regardless of type.
        if 'hidden' not in ext:
            continue
        ext_type = ext.get('type') or ''
        # …except types that carry the flag but aren't directory contacts (Sites),
        # where the value is meaningless and would read as a false "Hidden".
        if ext_type in NON_DIRECTORY_TYPES:
            continue
        hidden = bool(ext['hidden'])
        rows.append({
            'id': ext.get('id'),
            'extensionNumber': ext.get('extensionNumber') or '',
            'name': _display_name(ext),
            'type': ext_type or '—',
            'status': ext.get('status') or '',
            'hidden': hidden,
            'visible': not hidden,
            'editable': ext_type in USER_LIKE_TYPES,
        })

    # Sort by extension number (numeric where possible) for a stable directory view.
    def _sort_key(r):
        num = str(r.get('extensionNumber') or '')
        return (0, int(num)) if num.isdigit() else (1, num)
    rows.sort(key=_sort_key)

    visible = sum(1 for r in rows if r['visible'])
    editable = sum(1 for r in rows if r['editable'])
    summary = {
        'total': len(rows),
        'visible': visible,
        'hidden': len(rows) - visible,
        'editable': editable,
        'portal_only': len(rows) - editable,
    }
    return rows, summary


# ---------------------------------------------------------------------------
# Single-object update
# ---------------------------------------------------------------------------

def _error_message(resp):
    """Human-readable error string from an RC response, preserving RC's message."""
    if resp is None:
        return 'No response from RingCentral'
    try:
        body = resp.json() or {}
    except Exception:
        body = {}
    msg = body.get('message') or body.get('error')
    if not msg and isinstance(body.get('errors'), list) and body['errors']:
        msg = body['errors'][0].get('message')
    if not msg:
        msg = getattr(resp, 'text', '') or f"HTTP {getattr(resp, 'status_code', '?')}"
    return str(msg)[:300]


def set_visibility(ext_id, hidden, token):
    """Set an object's company-directory visibility.

    `hidden=True` hides it from the directory, `hidden=False` makes it visible.
    updateExtension accepts a partial body, so only the flag is sent. Returns
    (ok, message).
    """
    resp = rc_api_call(
        f"/restapi/v1.0/account/~/extension/{ext_id}",
        method='PUT', json={'hidden': bool(hidden)},
        token=token, return_response=True,
    )
    if resp is not None and getattr(resp, 'ok', False):
        return True, HIDDEN if hidden else VISIBLE
    return False, _error_message(resp)


# ---------------------------------------------------------------------------
# Bulk XLSX template
# ---------------------------------------------------------------------------

def generate_template(token):
    """Build the bulk-update workbook, pre-filled with the account's current
    directory-visibility state.

    Every supported object is written as a row with its current state and a
    ``New Visibility`` dropdown (Visible / Hidden) the operator edits. Rows left
    unchanged are skipped on apply.
    """
    rows, _summary = build_visibility_rows(token)
    if rows is None:
        raise Exception("Could not load account extensions. Token may be invalid or expired.")

    wb = Workbook()
    ws = wb.active
    ws.title = TEMPLATE_SHEET
    ws.append(TEMPLATE_HEADERS)

    for r in rows:
        current = HIDDEN if r['hidden'] else VISIBLE
        ws.append([
            r['id'], r['extensionNumber'], r['name'], r['type'],
            'Yes' if r['editable'] else 'No (portal only)',
            current, current,  # New Visibility defaults to current (no-op)
        ])

    # Reference sheet backing the New Visibility dropdown.
    ref = wb.create_sheet('Reference')
    ref['A1'] = 'Visibility'
    for i, v in enumerate(VISIBILITY_CHOICES, start=2):
        ref.cell(row=i, column=1, value=v)

    last = max(len(rows) + 1, 2)
    dv = DataValidation(
        type='list',
        formula1=f"=Reference!$A$2:$A${1 + len(VISIBILITY_CHOICES)}",
        allow_blank=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"G2:G{last}")  # New Visibility column (7th)

    widths = {'A': 14, 'B': 12, 'C': 30, 'D': 16, 'E': 18, 'F': 18, 'G': 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ref.column_dimensions['A'].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Cell cleaning
# ---------------------------------------------------------------------------

def _clean(value):
    """Normalise a spreadsheet cell to a trimmed string (handles NaN / '1001.0')."""
    text = str(value).strip()
    if text.lower() == 'nan':
        return ''
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    return text


def _parse_visibility(value):
    """Map a New Visibility cell to a hidden boolean.

    Accepts the dropdown values plus common synonyms. Returns True (hidden),
    False (visible), or None if the value isn't recognised.
    """
    v = _clean(value).lower()
    if v in ('visible', 'show', 'shown', 'yes', 'y', 'true'):
        return False
    if v in ('hidden', 'hide', 'not visible', 'no', 'n', 'false'):
        return True
    return None


# ---------------------------------------------------------------------------
# Bulk validate / apply (streamed NDJSON)
# ---------------------------------------------------------------------------

def process_upload_batch(records, token, is_preview=True, task_id=None):
    """Validate (preview) or apply (apply) directory-visibility changes from the
    uploaded rows, yielding NDJSON-friendly progress chunks:

        {"type": "start", ...}
        {"type": "progress", "result": {...}, "is_preview": bool}
        {"type": "cancelled", ...}   (apply only, on Stop)
        {"type": "done", "is_preview": bool}

    Each row is matched to a live object (by Extension ID, else Extension
    number). Rows whose New Visibility equals the object's current state are
    reported as no-ops and skipped; only genuine changes are applied.
    """
    total = len(records)
    yield {"type": "start", "total": total,
           "message": "Loading account directory (extensions)…"}

    rows, _summary = build_visibility_rows(token)
    if rows is None:
        yield {"type": "error", "message": "Could not load account extensions. Token may be invalid or expired."}
        return

    by_id = {str(r['id']): r for r in rows if r.get('id')}
    by_number = {}
    for r in rows:
        num = str(r.get('extensionNumber') or '').strip()
        if num:
            by_number.setdefault(num, r)

    changed = 0
    for i, row in enumerate(records):
        if not is_preview and task_control.is_stopped(task_id):
            yield {"type": "cancelled", "current": i, "total": total,
                   "message": f"Stopped by user. {i} of {total} row(s) processed; the rest were skipped."}
            task_control.clear(task_id)
            return

        ext_id = _clean(row.get('Extension ID', ''))
        ext_num = _clean(row.get('Extension', ''))
        new_vis_raw = row.get('New Visibility', '')

        def progress(status, message, target=None):
            return {
                "type": "progress",
                "current": i + 1,
                "total": total,
                "result": {
                    "row": i + 2,  # 1-based sheet row (header is row 1)
                    "ext": ext_num or (target or {}).get('extensionNumber') or "—",
                    "name": (target or {}).get('name') or "—",
                    "type": (target or {}).get('type') or "—",
                    "current": (VISIBLE if target and target['visible'] else HIDDEN) if target else "—",
                    "new": _clean(new_vis_raw) or "—",
                    "status": status,
                    "message": message,
                },
                "is_preview": is_preview,
            }

        # Fully blank row -> skip silently.
        if not any([ext_id, ext_num, _clean(new_vis_raw)]):
            yield progress("info", "Skipped empty row")
            continue

        # Resolve the target object (prefer the stable ID).
        target = by_id.get(ext_id) if ext_id else None
        if not target and ext_num:
            target = by_number.get(ext_num)
        if not target:
            key = ext_id or ext_num
            yield progress("error", f"No directory-visible object found for '{key}' on this account")
            continue

        desired_hidden = _parse_visibility(new_vis_raw)
        if desired_hidden is None:
            yield progress("error", f"Unrecognised New Visibility '{_clean(new_vis_raw)}' (use Visible or Hidden)")
            continue

        # No change requested -> nothing to do.
        if desired_hidden == target['hidden']:
            yield progress("info", f"Already {HIDDEN if desired_hidden else VISIBLE} — no change", target)
            continue

        # A change is requested, but only User-type objects can be changed via
        # the API — flag the rest instead of firing a doomed PUT.
        if not target.get('editable'):
            yield progress(
                "error",
                f"{target['type']} visibility can't be changed via the API "
                f"(User type only) — change it in the Admin Portal",
                target,
            )
            continue

        if is_preview:
            changed += 1
            arrow = f"{VISIBLE if target['visible'] else HIDDEN} → {HIDDEN if desired_hidden else VISIBLE}"
            yield progress("success", f"Will change {arrow}", target)
            continue

        # Apply mode -- flip the flag.
        ok, msg = set_visibility(target['id'], desired_hidden, token)
        if ok:
            changed += 1
            yield progress("success", f"Set to {msg}", target)
        else:
            yield progress("error", f"Update failed — {msg}", target)
        time.sleep(0.1)

    task_control.clear(task_id)
    yield {"type": "done", "is_preview": is_preview, "changed": changed}
