import io
import time
import pandas as pd
from openpyxl.styles import PatternFill, Font
from webapp.rc_api import rc_api_call
from webapp import task_control

# In-memory progress registry for the streamed audit job. Safe under the app's
# single-process, multi-threaded gunicorn deployment (see task_control.py for
# the same reasoning). Keyed by task_id.
audit_progress_store = {}

# Extension types that carry an editable ``regionalSettings.timezone`` and can
# therefore be audited against — and updated to match — the site they belong
# to. This covers people (users), call queues, and the other shared/system
# extensions RingCentral exposes a timezone for via the extension API.
#
# Single source of truth: each group is a (friendly label, [raw RC types])
# pair. The friendly label is what the audit shows and what the audit's
# "Extension Type" filter offers as a tick box; the raw types are the values
# RingCentral returns on an extension. Everything below is derived from this.
TYPE_FILTER_GROUPS = [
    ('User', ['User', 'DigitalUser', 'FlexibleUser']),
    ('Call Queue', ['Department']),
    ('Message Only', ['Voicemail']),
    ('Announcement Only', ['Announcement']),
    ('Site', ['Site']),
]

# RC extension types that represent a site itself. A site's timezone is its own
# regional setting, so it is compared against itself (never a false mismatch).
SITE_TYPES = {'Site'}

TIMEZONE_EXTENSION_TYPES = {t for _, raws in TYPE_FILTER_GROUPS for t in raws}

# Friendly label for each raw RC type. Anything unlisted falls back to itself.
EXTENSION_TYPE_LABELS = {t: label for label, raws in TYPE_FILTER_GROUPS for t in raws}

# Friendly label -> set of raw RC types, used to expand a selected filter label.
_LABEL_TO_TYPES = {label: set(raws) for label, raws in TYPE_FILTER_GROUPS}


def _type_label(raw_type):
    return EXTENSION_TYPE_LABELS.get(str(raw_type), str(raw_type or ''))


def type_filter_options():
    """Returns the ordered list of {'value','label'} extension-type tick boxes
    offered by the audit filter. Value and label are the same friendly string."""
    return [{'value': label, 'label': label} for label, _ in TYPE_FILTER_GROUPS]


def _expand_type_labels(labels):
    """Expands selected friendly type labels to the set of raw RC types. An
    empty/falsey selection means 'all auditable types'."""
    if not labels:
        return set(TIMEZONE_EXTENSION_TYPES)
    raws = set()
    for label in labels:
        raws |= _LABEL_TO_TYPES.get(str(label), set())
    return raws or set(TIMEZONE_EXTENSION_TYPES)


# The exact column order used for both the audit export and the update upload.
# Keeping them identical means the file the operator downloads is the same file
# they edit and upload back.
TEMPLATE_COLUMNS = [
    "Extension Name",
    "Extension Number",
    "Extension Type",
    "Site",
    "Site Timezone",
    "Extension Timezone",
]


# ---------------------------------------------------------------------------
# Directory fetching
# ---------------------------------------------------------------------------

def _paged(endpoint, token):
    """Yields every record across all pages of a collection endpoint."""
    page = 1
    while True:
        sep = "&" if "?" in endpoint else "?"
        resp = rc_api_call(
            f"{endpoint}{sep}perPage=1000&page={page}",
            token=token, raise_error=False
        )
        if not resp or 'records' not in resp:
            break
        for rec in resp['records']:
            yield rec
        if not resp.get('navigation', {}).get('nextPage'):
            break
        page += 1
        time.sleep(0.05)


# The alias used for the account's primary ("Main") site everywhere in this
# module, including as the filter value for the Main Site tick box.
MAIN_SITE_ID = 'main-site'


def _site_id_of(obj):
    """Effective site id for an extension/list record, defaulting to the Main
    Site alias when no site is set."""
    return str((obj.get('site') or {}).get('id', '')) or MAIN_SITE_ID


def list_sites(token):
    """Returns an ordered list of {'id','name'} for every site, including the
    Main Site alias. Used to populate the audit's Site filter tick boxes."""
    sites = [{'id': MAIN_SITE_ID, 'name': 'Main Site'}]
    for s in _paged("/restapi/v1.0/account/~/sites", token):
        sid = str(s.get('id'))
        if sid == MAIN_SITE_ID:
            continue
        sites.append({'id': sid, 'name': s.get('name', '') or sid})
    return sites


def fetch_timezone_extensions(token, types=None, sites=None):
    """Returns the list of extensions (basic list records) worth auditing.

    ``types`` is a set of raw RC extension types to include (defaults to every
    auditable type). ``sites`` is a set of site ids to include; an empty/None
    set means every site.
    """
    types = types or TIMEZONE_EXTENSION_TYPES
    site_set = set(sites) if sites else None
    out = []
    for e in _paged("/restapi/v1.0/account/~/extension", token):
        if e.get('type') not in types or not e.get('id'):
            continue
        # The Main Site has no extension number and is handled as a synthetic
        # row (see run_timezone_audit); skip the numberless site record here so
        # it isn't listed twice.
        if e.get('type') in SITE_TYPES and not str(e.get('extensionNumber', '')).strip():
            continue
        if site_set is not None and _site_id_of(e) not in site_set:
            continue
        out.append(e)
    return out


def resolve_main_site(token):
    """Returns (site_id, tz_name, tz_id) for the account's Main Site.

    The Main Site is identified in /account/~/sites by ``code == 'main-site'``
    (or the name 'Main Site'); its regional timezone falls back to the
    account-level regional settings when the site record omits it. ``site_id``
    is the real id when known, else the 'main-site' alias.
    """
    site_id = MAIN_SITE_ID
    tz = None
    for s in _paged("/restapi/v1.0/account/~/sites", token):
        is_main = (s.get('code') == 'main-site'
                   or str(s.get('name', '')).strip().lower() == 'main site')
        if is_main:
            site_id = str(s.get('id')) or MAIN_SITE_ID
            tz = (s.get('regionalSettings') or {}).get('timezone')
            if not tz and site_id != MAIN_SITE_ID:
                detail = rc_api_call(
                    f"/restapi/v1.0/account/~/sites/{site_id}",
                    token=token, raise_error=False
                )
                if detail:
                    tz = (detail.get('regionalSettings') or {}).get('timezone')
            break

    if not tz:
        account = rc_api_call("/restapi/v1.0/account/~", token=token, raise_error=False)
        tz = (account or {}).get('regionalSettings', {}).get('timezone')

    tz = tz or {}
    return site_id, str(tz.get('name', '')), str(tz.get('id', ''))


def build_site_timezone_map(token):
    """Returns {site_id (str): {'name': ..., 'tz_name': ...}}.

    The account's primary site is keyed under the well-known 'main-site' alias
    (and its real id, when known) using the account-level regional settings,
    so extensions on the Main Site can always be resolved.
    """
    site_map = {}
    for s in _paged("/restapi/v1.0/account/~/sites", token):
        sid = str(s.get('id'))
        tz = s.get('regionalSettings', {}).get('timezone')
        # The /sites collection does not always include regionalSettings, so
        # fetch the site detail when the timezone is missing.
        if not tz:
            detail = rc_api_call(
                f"/restapi/v1.0/account/~/sites/{sid}",
                token=token, raise_error=False
            )
            if detail:
                tz = detail.get('regionalSettings', {}).get('timezone')
            time.sleep(0.05)
        site_map[sid] = {
            'name': s.get('name', ''),
            'tz_name': str((tz or {}).get('name', '')),
        }

    # Account-level regional settings represent the primary ("Main") site,
    # which is not always returned by the /sites collection.
    account = rc_api_call("/restapi/v1.0/account/~", token=token, raise_error=False)
    if account:
        acc_tz = account.get('regionalSettings', {}).get('timezone')
        site_map.setdefault('main-site', {
            'name': 'Main Site',
            'tz_name': str((acc_tz or {}).get('name', '')),
        })

    return site_map


def _resolve_site(site_obj, site_map):
    """Maps an extension's site object to (site_name, site_tz_name)."""
    site_obj = site_obj or {}
    site_id = str(site_obj.get('id', '')) or 'main-site'
    info = site_map.get(site_id)
    if info is None and site_id != 'main-site':
        info = site_map.get('main-site')
    info = info or {}
    name = info.get('name') or site_obj.get('name', '') or 'Main Site'
    return name, info.get('tz_name', '')


# ---------------------------------------------------------------------------
# Audit (streamed background job)
# ---------------------------------------------------------------------------

def _build_audit_workbook(rows):
    """Renders audit rows to an .xlsx BytesIO, mismatches sorted to the top and
    highlighted in red."""
    df = pd.DataFrame(rows, columns=TEMPLATE_COLUMNS)

    if not df.empty:
        mismatch = df["Site Timezone"].astype(str).str.strip() != df["Extension Timezone"].astype(str).str.strip()
        df["_sort"] = (~mismatch).astype(int)  # mismatches (False->0) first
        df.sort_values(by=["_sort", "Site", "Extension Number"], inplace=True)
        df.drop(columns=["_sort"], inplace=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Extension Timezone Audit')
        ws = writer.sheets['Extension Timezone Audit']

        for column in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(length + 5, 50)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        headers = [c.value for c in ws[1]]
        site_col = headers.index("Site Timezone") + 1
        ext_col = headers.index("Extension Timezone") + 1
        for row in range(2, ws.max_row + 1):
            site_tz = str(ws.cell(row=row, column=site_col).value or "").strip()
            ext_tz = str(ws.cell(row=row, column=ext_col).value or "").strip()
            if site_tz != ext_tz:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red

    output.seek(0)
    return output.getvalue()


def run_timezone_audit(task_id, token, type_labels=None, site_ids=None):
    """Background worker: builds the per-extension timezone-vs-site audit and
    stores the resulting workbook in ``audit_progress_store`` for download.

    ``type_labels`` and ``site_ids`` are the audit filter selections (friendly
    extension-type labels and site ids). An empty selection means 'all'.
    """
    audit_progress_store[task_id] = {
        'current': 0, 'total': 0, 'status': 'running', 'file_ready': False, 'mismatches': 0
    }
    try:
        site_map = build_site_timezone_map(token)
        extensions = fetch_timezone_extensions(
            token,
            types=_expand_type_labels(type_labels),
            sites=set(site_ids) if site_ids else None,
        )

        store = audit_progress_store[task_id]
        store['total'] = len(extensions)

        rows = []
        mismatches = 0
        for idx, ext in enumerate(extensions):
            if task_control.is_stopped(task_id):
                store['status'] = 'cancelled'
                task_control.clear(task_id)
                return

            detail = rc_api_call(
                f"/restapi/v1.0/account/~/extension/{ext['id']}",
                token=token, raise_error=False
            )
            time.sleep(0.05)
            store['current'] = idx + 1

            if not detail:
                continue

            ext_tz_name = str(detail.get('regionalSettings', {}).get('timezone', {}).get('name', ''))
            ext_type = detail.get('type') or ext.get('type')

            if ext_type in SITE_TYPES:
                # A site's timezone is its own regional setting — it is its own
                # "site", so it is compared against itself (never a mismatch).
                site_name = detail.get('name', '')
                site_tz_name = ext_tz_name
            else:
                site_name, site_tz_name = _resolve_site(detail.get('site'), site_map)

            if ext_tz_name.strip() != site_tz_name.strip():
                mismatches += 1

            rows.append({
                "Extension Name": detail.get('name', ''),
                "Extension Number": str(detail.get('extensionNumber', '')),
                "Extension Type": _type_label(ext_type),
                "Site": site_name,
                "Site Timezone": site_tz_name,
                "Extension Timezone": ext_tz_name,
            })

        # The Main Site has no extension number, so it isn't returned by
        # /extension. Add it as a synthetic Site row when the Site type and the
        # Main Site both pass the audit filters. Its Extension Number carries
        # the 'main-site' sentinel so it can be targeted on upload.
        label_set = set(type_labels or [])
        site_selected = (not label_set) or ('Site' in label_set)
        main_selected = (not site_ids) or (MAIN_SITE_ID in set(site_ids))
        if site_selected and main_selected:
            _mid, main_tz, _tzid = resolve_main_site(token)
            rows.append({
                "Extension Name": "Main Site",
                "Extension Number": MAIN_SITE_ID,
                "Extension Type": _type_label('Site'),
                "Site": "Main Site",
                "Site Timezone": main_tz,
                "Extension Timezone": main_tz,
            })

        store['mismatches'] = mismatches
        store['file_data'] = _build_audit_workbook(rows)
        store['status'] = 'completed'
        store['file_ready'] = True
    except Exception as e:
        audit_progress_store[task_id]['status'] = 'error'
        audit_progress_store[task_id]['error'] = str(e)


# ---------------------------------------------------------------------------
# Update (streamed upload -> preview / apply)
# ---------------------------------------------------------------------------

def _load_timezone_dictionary(token):
    """Returns (records, name_to_id) for the RC timezone dictionary."""
    records = list(_paged("/restapi/v1.0/dictionary/timezone", token))
    name_to_id = {}
    for tz in records:
        name_to_id[str(tz.get('name', '')).lower().strip()] = str(tz.get('id'))
        name_to_id[str(tz.get('id'))] = str(tz.get('id'))
    return records, name_to_id


def _resolve_timezone_id(raw, name_to_id, records):
    """Resolves a user-supplied timezone string to a RC timezone id, tolerant of
    the canonical name ('Australia/Sydney'), the id, or a city/description."""
    raw_lower = str(raw).lower().strip()
    if not raw_lower:
        return None
    if raw_lower in name_to_id:
        return name_to_id[raw_lower]
    city = raw_lower.split('/')[-1].replace('_', ' ')
    for tz in records:
        if city and (city in str(tz.get('name', '')).lower()
                     or city in str(tz.get('description', '')).lower()):
            return str(tz['id'])
    return None


def _get(row, *keys):
    """Returns the first non-empty value across the given column names.

    Multiple keys support backward compatibility with files exported by the
    older 'User Timezone Audit' tool (e.g. 'User Timezone' / 'User Name').
    """
    for key in keys:
        val = row.get(key, '')
        if val is None:
            continue
        val = str(val).strip()
        if val:
            return val
    return ''


def update_timezone_batch(records, token, is_preview=False, task_id=None):
    """Generator yielding NDJSON-friendly progress dicts while previewing or
    applying extension timezone changes from an uploaded audit file."""
    total = len(records)
    yield {"type": "start", "total": total,
           "message": "Loading account directories..."}

    # Verify the token and load supporting directories.
    account = rc_api_call("/restapi/v1.0/account/~", token=token, raise_error=False)
    if not account:
        yield {"type": "error", "message": "Unauthorized. Token expired or invalid."}
        return

    tz_records, name_to_id = _load_timezone_dictionary(token)
    if not tz_records:
        yield {"type": "error", "message": "Failed to load the timezone dictionary."}
        return

    # Build an extension-number -> id map across every extension in the account.
    ext_map = {}
    for e in _paged("/restapi/v1.0/account/~/extension", token):
        num = str(e.get('extensionNumber', '')).strip()
        if num and e.get('id'):
            ext_map[num] = str(e['id'])

    updated = skipped = errors = 0

    for i, row in enumerate(records):
        if not is_preview and task_control.is_stopped(task_id):
            yield {"type": "cancelled", "current": i, "total": total,
                   "message": f"Stopped by user. {i} of {total} row(s) processed; the rest were skipped."}
            task_control.clear(task_id)
            return

        name = _get(row, "Extension Name", "User Name")
        ext_num = _get(row, "Extension Number").split('.')[0]
        desired_raw = _get(row, "Extension Timezone", "User Timezone")

        def emit(status, message, from_tz="", to_tz=""):
            return {
                "type": "row", "current": i + 1, "total": total,
                "is_preview": is_preview, "name": name, "ext": ext_num,
                "from": from_tz, "to": to_tz, "status": status, "message": message,
            }

        if not ext_num:
            errors += 1
            yield emit("error", "Missing Extension Number.")
            continue

        # The Main Site has no extension number; it carries the 'main-site'
        # sentinel and is updated through the Sites API rather than /extension.
        if ext_num.lower() == MAIN_SITE_ID:
            new_tz_id = _resolve_timezone_id(desired_raw, name_to_id, tz_records)
            if not new_tz_id:
                errors += 1
                yield emit("error", f"Unrecognised timezone: '{desired_raw}'.")
                continue
            new_tz_name = next((str(t.get('name', '')) for t in tz_records
                                if str(t.get('id')) == new_tz_id), desired_raw)
            site_id, cur_tz_name, cur_tz_id = resolve_main_site(token)

            if cur_tz_id == new_tz_id:
                skipped += 1
                yield emit("skipped", "Already set — no change.", cur_tz_name, new_tz_name)
                continue
            if is_preview:
                yield emit("pending", "Will update.", cur_tz_name, new_tz_name)
                continue

            payload = {"regionalSettings": {"timezone": {"id": new_tz_id}}}
            resp = rc_api_call(
                f"/restapi/v1.0/account/~/sites/{site_id}",
                method="PUT", json=payload, token=token, return_response=True
            )
            time.sleep(0.05)
            if resp is not None and getattr(resp, 'ok', False):
                updated += 1
                yield emit("applied", "Timezone updated.", cur_tz_name, new_tz_name)
            else:
                errors += 1
                err_text = ""
                try:
                    err_text = str(resp.json().get('message', '')) if resp is not None else ""
                except Exception:
                    err_text = getattr(resp, 'text', '') if resp is not None else ""
                yield emit("error", f"Update failed. {err_text}".strip(), cur_tz_name, new_tz_name)
            continue

        ext_id = ext_map.get(ext_num)
        if not ext_id:
            errors += 1
            yield emit("error", f"No extension found for number {ext_num}.")
            continue

        new_tz_id = _resolve_timezone_id(desired_raw, name_to_id, tz_records)
        if not new_tz_id:
            errors += 1
            yield emit("error", f"Unrecognised timezone: '{desired_raw}'.")
            continue

        detail = rc_api_call(
            f"/restapi/v1.0/account/~/extension/{ext_id}",
            token=token, raise_error=False
        )
        if not detail:
            errors += 1
            yield emit("error", "Failed to read the current extension settings.")
            continue

        cur_tz = detail.get('regionalSettings', {}).get('timezone', {}) or {}
        cur_tz_id = str(cur_tz.get('id', ''))
        cur_tz_name = str(cur_tz.get('name', ''))
        new_tz_name = next((str(t.get('name', '')) for t in tz_records
                            if str(t.get('id')) == new_tz_id), desired_raw)

        if cur_tz_id == new_tz_id:
            skipped += 1
            yield emit("skipped", "Already set — no change.", cur_tz_name, new_tz_name)
            continue

        if is_preview:
            yield emit("pending", "Will update.", cur_tz_name, new_tz_name)
            continue

        payload = {"regionalSettings": {"timezone": {"id": new_tz_id}}}
        resp = rc_api_call(
            f"/restapi/v1.0/account/~/extension/{ext_id}",
            method="PUT", json=payload, token=token, return_response=True
        )
        time.sleep(0.05)
        if resp is not None and getattr(resp, 'ok', False):
            updated += 1
            yield emit("applied", "Timezone updated.", cur_tz_name, new_tz_name)
        else:
            errors += 1
            err_text = ""
            try:
                err_text = str(resp.json().get('message', '')) if resp is not None else ""
            except Exception:
                err_text = getattr(resp, 'text', '') if resp is not None else ""
            yield emit("error", f"Update failed. {err_text}".strip(), cur_tz_name, new_tz_name)

    if not is_preview:
        task_control.clear(task_id)

    yield {"type": "done", "is_preview": is_preview,
           "updated": updated, "skipped": skipped, "errors": errors,
           "total": total}
